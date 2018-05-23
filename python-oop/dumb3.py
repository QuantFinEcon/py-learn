####################################
# compare tables
####################################

#####################################
# LISTEN_PYTHON.PY
#####################################


import win32com.client
import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil

#os.chdir('C:\\Users\\1580873\\Desktop\\RiskView')
import outlook_helpers
#from importlib import reload
#reload(outlook_helpers)

#==============================================================================
# handler class for incoming mail (FOR pythoncom LISTENING)
#==============================================================================

class Handler_Class(object):
    
    # mutate class attr without initialising class
    _IC = ""
    
    def __init__(self, email=None):
        self._outlook = win32com.client.Dispatch('outlook.application')
        self._attachment = None
        self._GAL = outlook_helpers.outlookGAL()
        self._results = {}
    
    def OnNewMailEx(self, receivedItemsIDs):
        # https://msdn.microsoft.com/en-us/vba/outlook-vba/articles/application-newmailex-event-outlook
        for ID in receivedItemsIDs.split(","):
            mailItem = self._outlook.Session.GetItemFromID(ID)
            
            # ensure mail is incoming to the relevant inbox            
            try:
                self._GAL.find(mailItem.To)
                to_mail = self._GAL.results[mailItem.To]['email']
                self._GAL.del_results()                
            except Exception as error: 
                print("Not a recognised email in Global Address List")
                print( str(error) )
                return
            if to_mail != self._IC: 
                print("Not to relevant inbox!")
                return
            else:
                print("To relevant inbox!")
                pass
            
            # multiple mails incoming simultaneously
            print("Subject: " + mailItem.Subject)
            print("Body: " + mailItem.Body)
            print("Sender: " + mailItem.SenderName)
#            print("att counts" + str(mailItem.Attachments.Count))
#            print('attach:' + mailItem.Attachments.Item(1).DisplayName)

            if 'autoreply' in mailItem.Subject.lower():
                print("Don't reply an auto reply...")
                return
            elif 'riskview' in mailItem.Subject.lower(): request_type = 'RiskView'
            elif 'ccbr' in mailItem.Subject.lower(): request_type = 'CCBR'
            #elif 'mtcr' in mailItem.Subject.lower(): request_type = 'MTCR'
            else: 
                print('Not a mail to reply... No action!')
                return
            
            switcher = {
                    'RiskView': self._OnNewMailEx_RiskView,
                    'CCBR': self._OnNewMailEx_CCBRView #,'MTCR': self._OnNewMailEx_MTCRView
                        }
    
            # case: call private method
            process = switcher.get(request_type, lambda: "invalid mail checker rule!")
            process(mailItem)
            print("===============================")
    
    def _OnNewMailEx_RiskView(self, mailItem):
        
        check = self._OnNewMail_RiskView_filter(mailItem)
        if check is True:
            # edit data into exisiting file                
            # key --> RequestType|BankID|Name|Department
            fields = list(self._results.keys())[0].split('|')
            
            record = pd.DataFrame(fields[1:]+[None]).transpose()
            record.columns = ['BankID','Name','UserGroup','Dummy']
            
            rw_subsegmentaccess = self._results[list(self._results.keys())[0]]['rw_subsegmentaccess']
            subsegmentcode = rw_subsegmentaccess.loc[rw_subsegmentaccess['Grant Access']=='Yes','SubSegment Code'].to_frame()                
            subsegmentcode['Dummy'] = None
            subsegmentcode.columns = ['Subsegment','Dummy']
            
            rw_countryaccess = self._results[list(self._results.keys())[0]]['rw_countryaccess']
            countrycode = rw_countryaccess.loc[rw_countryaccess['Grant Access']=='Yes','Country Code'].to_frame()                
            countrycode['Dummy'] = None
            countrycode.columns = ['CountryCode','Dummy']
            
            all_CB = all([ x == 'Yes' for x in rw_subsegmentaccess.loc[rw_subsegmentaccess['CB / CIB']=='CB','Grant Access'] ])
            all_CIB = all([ x == 'Yes' for x in rw_subsegmentaccess.loc[rw_subsegmentaccess['CB / CIB']=='CIB','Grant Access'] ])
            all_Subsegment = all([ x == 'Yes' for x in rw_subsegmentaccess.loc[:,'Grant Access'] ])
            global_Countries = all([ x == 'Yes' for x in rw_countryaccess.loc[:,'Grant Access'] ])

            if global_Countries and \
                (all_CB or all_CIB or all_Subsegment):
                # Global + CB / Global + CIB / Global + all subsegment
                # copy to DASHBOARD access only, not for RiskView Access / CCBRView Access
                # can keep a seperate internal sheet indpt from Tableau for recordkeeping of global access users
                d_tocopy = record.drop('Dummy',axis=1)
                d_tocopy['Project Access'] = 'RiskView'
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'DashboardAccess', \
                        datatoappend = d_tocopy)
                
                d_tocopy = d_tocopy.drop('Project Access',axis=1)
                
                if all_Subsegment:                    
                    d_tocopy['Global Type'] = "Global/Global"
                elif all_CB:
                    d_tocopy['Global Type'] = "Global/CB"
                elif all_CIB:
                    d_tocopy['Global Type'] = "Global/CIB"
                
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'RiskViewGlobalViewAccess', \
                        datatoappend = d_tocopy)
                
            else:
                #custom countries
                # CROSS APPLY if custom
                tocopy = pd.merge(record, subsegmentcode, on='Dummy')
                tocopy = pd.merge(tocopy, countrycode, on='Dummy').drop('Dummy',axis=1)
                # copy and paste into results or data of form into Pending_Access.xlsx Stack
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                                        sheetname = 'RiskViewAccess', \
                                        datatoappend = tocopy)
                d_tocopy = record.drop('Dummy',axis=1)
                d_tocopy['Project Access'] = 'RiskView'
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'DashboardAccess', \
                        datatoappend = d_tocopy)
            
            # save mailItem as .msg into cwd\\folder            
            if not os.path.isdir(os.getcwd() + '\\' + 'RiskView_MailItems'):
                os.mkdir('RiskView_MailItems')
            msgName = 'RiskView Request ' + fields[0]
            mailItem.SaveAs(os.getcwd() + '\\' + 'RiskView_MailItems\\' + msgName + '.msg')

        else:
            # failed filter, auto reply rejection
            to = mailItem.SenderName
            subject = "AUTOREPLY: " + mailItem.Subject
            body = self._rejection_pretty(check)
            self._send_mail(to, subject, body, mailItem) # handles 0 or multiple attachments

        pass
        
    def _OnNewMailEx_CCBRView(self, mailItem):
        
        check = self._OnNewMail_CCBRView_filter(mailItem, approver_department = 'Group Country Risk')
        if check is True:
            # edit data into exisiting file                
            # key --> RequestType|BankID|Name|Department
            fields = list(self._results.keys())[0].split('|')
            
            record = pd.DataFrame(fields[1:]+[None]).transpose()
            record.columns = ['BankID','Name','UserGroup','Dummy']
            
            rw_countryaccess = self._results[list(self._results.keys())[0]]['rw_countryaccess']
            countrycode = rw_countryaccess.loc[rw_countryaccess['Grant Access']=='Yes','Country Code'].to_frame()                
            countrycode['Dummy'] = None
            countrycode.columns = ['CountryCode','Dummy']
            
            global_Countries = all([ x == 'Yes' for x in rw_countryaccess.loc[:,'Grant Access'] ])

            if global_Countries:
                # Global + CB / Global + CIB / Global + all subsegment
                # copy to DASHBOARD access only, not for RiskView Access / CCBRView Access
                # can keep a seperate internal sheet indpt from Tableau for recordkeeping of global access users
                d_tocopy = record.drop('Dummy',axis=1)
                d_tocopy['Project Access'] = 'CCBRView'
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'DashboardAccess', \
                        datatoappend = d_tocopy)

                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'CCBRViewGlobalViewAccess', \
                        datatoappend = d_tocopy.drop('Project Access',axis=1))
                
            else:
                #custom countries
                # CROSS APPLY if custom
                tocopy = pd.merge(record, countrycode, on='Dummy').drop('Dummy',axis=1)
                # copy and paste into results or data of form into Pending_Access.xlsx Stack
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                                        sheetname = 'CCBRViewAccess', \
                                        datatoappend = tocopy)
                d_tocopy = record.drop('Dummy',axis=1)
                d_tocopy['Project Access'] = 'CCBRView'
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'DashboardAccess', \
                        datatoappend = d_tocopy)
            
            # save mailItem as .msg into cwd\\folder            
            if not os.path.isdir(os.getcwd() + '\\' + 'RiskView_MailItems'):
                os.mkdir('RiskView_MailItems')
            msgName = 'RiskView Request ' + fields[0]
            mailItem.SaveAs(os.getcwd() + '\\' + 'RiskView_MailItems\\' + msgName + '.msg')
            
        else:
            # failed filter, auto reply rejection
            to = mailItem.SenderName
            subject = "AUTOREPLY: " + mailItem.Subject
            body = self._rejection_pretty(check)
            self._send_mail(to, subject, body, mailItem) # handles 0 or multiple attachments

        pass
        
    def _OnNewMailEx_MTCRView(self, mailItem):
        
        check = self._OnNewMail_MTCRView_filter(mailItem)
        if check is True:
            pass
        else:
            # failed filter, auto reply rejection
            pass
    
    def _append_to_pending(self, filename, sheetname, datatoappend, index=False, header=False):
        #with open('Pending_Access.xlsx', 'a') as f:
            #    df.to_csv(f, header=True, index_label=False)
            
        # http://openpyxl.readthedocs.io/en/default/pandas.html
        
        # from openpyxl import load_workbook
        # from openpyxl.utils.dataframe import dataframe_to_rows            
        wb = load_workbook(filename = filename)
        #wb.sheetnames
        sht = wb[sheetname]
        #sht.max_row
        for row in dataframe_to_rows(datatoappend, index=index, header=header):
            sht.append(row)

        wb.save(filename)
        wb.close()
        del(row)
        del(sht)
        del(wb)        
        pass
    
    
    def _OnNewMail_CCBRView_filter(self, mailItem, *args, **kwargs):
        '''
        *args, **kwargs
        approver_department = "Group Country Risk"
        '''
        # CORRECT EMAIL SUBJECT
        rule1 = 'request' or 'ccbr' in str(mailItem.Subject).lower()
        if not rule1: return "Mail subject incorrectly named.\n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        # HAS ATTACHMENT AND ONLY RELEVANT ONE ATTACHMENT
        rule2 = mailItem.Attachments.Count == 1
        if not rule2: return "Mail do not have the request form attachment and can only have one attachment.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION
        
        # CHECK FILENAME OF ATTACHMENT FORM
        rule3 = mailItem.Attachments.Item(1).DisplayName == "CCBRView Access Request Form.xlsx"
        if not rule3: return "Mail do not have the correct request form attachment.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION


        if not os.path.isdir(os.getcwd() + '\\' + 'temporary_for_attachments'):
            os.mkdir('temporary_for_attachments')
        old_dir = os.getcwd()
        os.chdir(os.getcwd() + '\\' + 'temporary_for_attachments')
        # OPEN ATTACHMENT
        self._open_attachments(mailItem, whichItem=1, sheetname = 0, 
                               index_col=None, header=None, savein = os.getcwd())
        os.chdir(old_dir)
        #os.rmdir('temporary_for_attachments')
        
        # TALLY BankID TO ENSURE macros was used to send request form
        rule4 = len(re.findall(r"(?<!\d)\d{7}(?!\d)", mailItem.Subject)) == 1
        if not rule4: return "Request form can only have one requester's BankID.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION

        requester_BankID = str(self._attachment.iloc[0,1])
        rule5 = requester_BankID == str(re.findall(r"(?<!\d)\d{7}(?!\d)", mailItem.Subject)[0])
        if not rule5: return "Please include requester's BankID in mail subject.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION
        
        # VALIDATE DATE ACCURACY AND COMPLETENESS IN REQUEST FORM
        rw_countryaccess = self._attachment.iloc[4:265, 0:4]
        rw_countryaccess.columns = rw_countryaccess.iloc[0]
        rw_countryaccess = rw_countryaccess.reindex(rw_countryaccess.index.drop(4))
        rule6 = rw_countryaccess.shape == (260, 4) and \
                all(entry in ['Yes','No'] for entry in rw_countryaccess['Grant Access'])
        if not rule6: return "Request form attachment country access data is invalidated.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION
        
        # So far, SENDER is anyone from Department Group Country Risk
        # requester --> line manager --> our inbox
        
        self._GAL.find(requester_BankID)
        try:
            rule8 = self._GAL.results[requester_BankID]["Department"] == kwargs['approver_department']
        except:
            print('Key error: CCBRView filter needs an approver department from ' + kwargs['approver_department'])
            return        
        
        self._GAL.del_results()
        
        if not rule8: return "Warning: Approver for CCBRView should be from the Group Country Risk Department.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION
                
        # passed all rejection filters
        # key --> RequestType|BankID|Name|Department
        self._GAL.find(requester_BankID)
        name = self._GAL.get_results()['Name'].iloc[0]
        dept = self._GAL.get_results()['Department'].iloc[0]
        self._GAL.del_results()
        
        self._results["CCBRView|"+str(requester_BankID)+"|"+str(name)+"|"+str(dept)] = \
                        {"rw_countryaccess":rw_countryaccess,
                         "Business Justification":self._attachment.iloc[1,1]}
        return True

    def _OnNewMail_MTCRView_filter(self, mailItem):
        # So far, SENDER is Head of MTCR # approver = "Poquet, Morgan"
        # requester --> line manager --> our inbox
        approver = mailItem.SenderName
        self._GAL.find(approver)
        rule1 = self._GAL.results[approver]["JobTitle"] in \
            ["Head, MTCR - Traded Credit & Credit Trading", "Global Head Market & Traded Credit Risk"]
        self._GAL.del_results()
        
        # Alternatively, SENDER is Head of MTCR is part of FWD mail thread
        # requester --> line manager --> requester FWD --> our inbox
        if not rule1:
            single_mail_body = self._get_single_email_thread(mailItem.Body, approver)
            mail_list = [x.strip(',.-><:;') for x in single_mail_body.lower().split()]
            matchingkw = ['approve']
            rule1 = any(x in matchingkw for x in mail_list)
        
        if not rule1: return "Please be informed that only the Head of MTCR can approve request access.\n\
        Please get approval from the Head of MTCR and resubmit." # proceed to AUTO REJECTION    
        
        # passed all rejection filters
#        self._results["MTCRView|"+str(requester_BankID)+"|"+str(name)+"|"+str(dept)] = \
#                        {"rw_countryaccess":rw_countryaccess,
#                         "Business Justification":self._attachment.iloc[1,1]}
        return True
        
    def _OnNewMail_RiskView_filter(self, mailItem, *args, **kwargs):
        '''
        *args, **kwargs
        matchingkw = ['approve']
        '''
        # CORRECT EMAIL SUBJECT
        rule1 = 'request' and 'riskview' in str(mailItem.Subject).lower()
        if not rule1: return "Mail subject incorrectly named. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        # HAS ATTACHMENT AND ONLY RELEVANT ONE ATTACHMENT
        rule2 = mailItem.Attachments.Count == 1
        if not rule2: return "Mail do not have the request form attachment and can only have one attachment. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        # CHECK FILENAME OF ATTACHMENT FORM
        rule3 = mailItem.Attachments.Item(1).DisplayName == "RiskView Access Request Form.xlsm"
        if not rule3: return "Mail do not have the correct request form attachment. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        if not os.path.isdir(os.getcwd() + '\\' + 'temporary_for_attachments'):
            os.mkdir('temporary_for_attachments')
        old_dir = os.getcwd()
        print(old_dir)
        os.chdir(os.getcwd() + '\\' + 'temporary_for_attachments')
        # OPEN ATTACHMENT
        self._open_attachments(mailItem, whichItem=1, sheetname = 0, 
                               index_col=None, header=None, savein = os.getcwd())
        os.chdir(old_dir)
        shutil.rmtree(os.getcwd() + '//' + 'temporary_for_attachments')
        
        # TALLY BankID TO ENSURE macros was used to send request form
        rule4 = len(re.findall(r"(?<!\d)\d{7}(?!\d)", mailItem.Subject)) == 1
        if not rule4: return "Request form can only have one requester's BankID. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION        

        requester_BankID = str(self._attachment.iloc[0,1])
        rule5 = requester_BankID == str(re.findall(r"(?<!\d)\d{7}(?!\d)", mailItem.Subject)[0])
        if not rule5: return "Please include requester's BankID in mail subject.\n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        # VALIDATE DATE ACCURACY AND COMPLETENESS IN REQUEST FORM
        rw_countryaccess = self._attachment.iloc[4:263, 0:4]
        rw_countryaccess.columns = rw_countryaccess.iloc[0]
        rw_countryaccess = rw_countryaccess.reindex(rw_countryaccess.index.drop(4))
        rule6 = rw_countryaccess.shape == (258, 4) and \
                all(entry in ['Yes','No'] for entry in rw_countryaccess['Grant Access'])
        if not rule6: return "Request form attachment country access data is invalidated. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        rw_subsegmentaccess = self._attachment.iloc[4:48, 5:11]
        rw_subsegmentaccess.columns = rw_subsegmentaccess.iloc[0]
        rw_subsegmentaccess = rw_subsegmentaccess.reindex(rw_subsegmentaccess.index.drop(4))
        rule7 = rw_subsegmentaccess.shape == (43, 6) and \
                all(entry in ['Yes','No'] for entry in rw_subsegmentaccess['Grant Access'])
        if not rule7: return "Request form attachment subsegment access data is invalidated. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION

        # CHECK APPROVER who send the email is the requester's LINE MANAGER
        # requester --> line manager --> our inbox
        self._GAL.find_manager(requester_BankID)
        if len(self._GAL.results) == 1: # one worker --> one line manager
            approver_BankID = list(self._GAL.results.keys())[0]
            approver_name = self._GAL.results[approver_BankID]['Name']
            sendername = mailItem.SenderName # line manager needs to send the mail via macros
            rule8 = sendername == approver_name
        else:
            rule8 = False # no manager, invalid BankID
        self._GAL.del_results()
        
        # ALTERNATIVE CHECK OF APPROVER (those who did not use the macros)
        # requester --> line manager --> requester FWD --> our inbox
        if not rule8:
            single_mail_body = self._get_single_email_thread(mailItem.Body, approver_name)
            mail_list = [x.lower().strip(',.-><:;') for x in single_mail_body]
            matchingkw = ['approve']
            rule8 = any(x in matchingkw for x in mail_list)

        if not rule8: return "Warning: Approver is not the requester's direct line manager. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION          
        
        # passed all rejection filters
        # key --> RequestType|BankID|Name|Department
        self._GAL.find(requester_BankID)
        name = self._GAL.get_results()['Name'].iloc[0]
        dept = self._GAL.get_results()['Department'].iloc[0]
        self._GAL.del_results()
        
        self._results["RiskView|"+str(requester_BankID)+"|"+str(name)+"|"+str(dept)] = \
                        {"rw_countryaccess":rw_countryaccess,
                         "rw_subsegmentaccess":rw_subsegmentaccess,
                         "Business Justification":self._attachment.iloc[1,1]}
        
        return True

    def _get_single_email_thread(self, mailItemBody, from_who):
        # If in Body: Must be from: <Line Manager Name> and below it includes Approved before next From
        # found in: SenderName, Subject, Body, 
        # extra mail Attributes: ReceivedTime, Attachments
        
        pattern = "From: " + from_who
        single_mail = mailItemBody[mailItemBody.find(pattern):]
        ending = single_mail[len(pattern):].find('From: ')
        ending = None if ending == -1 else ending + len(pattern)
        
        if ending is None:
            if single_mail == '\n': return []
            return [single_mail]
        else:
            return [single_mail[ :ending]] + \
                self._get_single_email_thread(self, single_mail[ending: ], from_who)


    def _rejection_pretty(self, msg):
        # rejection auto reply
        return "================================" + "\n" + \
        " ROBOT AUTO REJECTION " + "\n" + \
        "================================" + "\n" + \
        msg + "\n" + "\n" + \
        "Please resubmit with the relevant attachment. Thanks for your cooperation!" + "\n" + \
        "\n" + \
        "From the Credit Risk Monitoring Team"

    def _save_attachments(self, mailItem, whichItem=1, savein = os.getcwd()):
        '''
        args=['io', 'sheetname', 'header', 'skiprows', 'skip_footer', 'index_col', 
        'names', 'parse_cols', 'parse_dates', 'date_parser', 'na_values', 'thousands', 
        'convert_float', 'has_index_names', 'converters', 'dtype', 'true_values', 
        'false_values', 'engine', 'squeeze'], varargs=None, keywords='kwds', 
        '''
        # .xlsm means cannot copy paste into another excel wb
        # Don't accept i.e. Copy of RiskView Access Request Form.xlsm
        filename = mailItem.Attachments.Item(whichItem).DisplayName
        # save current attachment as temp in local drive
        print("saving attachments in... " + savein + '\\' + str(filename))
        mailItem.Attachments.Item(whichItem).SaveAsFile(savein + '\\' + str(filename))
        return savein + '\\' + str(filename)

    def _open_attachments(self, mailItem, whichItem=1, savein = os.getcwd(), 
                          *args, **kwargs):
        filename = self._save_attachments(mailItem, whichItem, savein)
        # read from local drive for processing
        self._attachment = pd.read_excel(io = filename, *args, **kwargs)
        pass
    
    def _send_mail(self, to, subject=None, body=None, attachments=None):
        mail = self._outlook.CreateItem(0)
        mail.To = to
        #Msg.CC = "more email addresses here"
        #Msg.BCC = "more email addresses here"
        if subject is not None: mail.Subject = subject
        if body is not None: mail.Body = body
        #In case you want to attach a file to the email
        #attachment  = "C:\\Users\\1580873\\Desktop\\IFRS9\\Script\\anthony_stdf_mapping.R"
        
        if not os.path.isdir(os.getcwd() + '\\' + 'temporary_for_attachments'):
            os.mkdir('temporary_for_attachments')
        old_dir = os.getcwd()
        os.chdir(os.getcwd() + '\\' + 'temporary_for_attachments')
        if attachments is not None:
            if isinstance(attachments,str): 
                mail.Attachments.Add(attachments)
            # handle multiple attachments ['C://...', 'C://...']
            elif isinstance(attachments,list): 
                for att in attachments:
                    mail.Attachments.Add(att)
            # forward attachments from another Outlook.mailItem
            elif 'win32com.gen_py' and 'MailItem' in str(type(attachments)):
                for i in range(1,1+attachments.Attachments.Count):
                    filename = self._save_attachments(attachments, whichItem = i, savein = os.getcwd())
                    mail.Attachments.Add(filename)
                    os.remove(filename)
            else:
                attachments = None
        mail.Send()
        os.chdir(old_dir)
        shutil.rmtree(os.getcwd() + '//' + 'temporary_for_attachments')
        pass
    
    
    __doc__ = """
    ====================================================================
    USES
    ====================================================================
    - print string from python to log.txt saved on local drive
    ====================================================================    
    SAMPLE
    ====================================================================
    log = FileWriter("filename.txt")
    log.printlog("some string to log into .txt file")
    log.flush()
    """            



#####################################
# OUTLOOK_HELPERS.PY
#####################################



import win32com.client
#from win32com.client.gencache import EnsureDispatch as Dispatch
import pandas as pd
import re
import os

#==============================================================================
# Read current Window's Outlook Global Address List
#==============================================================================

class outlookGAL(object):

    def __init__(self):
        self._outlook = win32com.client.gencache.EnsureDispatch('Outlook.Application')
        self._BankID = ''
        self.results = {}

    def _getfromOutlook(self, BankID, isManager = False, workerBankID = None):
        
        recipient = self._outlook.Session.CreateRecipient(BankID)
        rev = recipient.Resolve()
        if not rev:
            print("BankID: " + str(BankID) + " is not found in Global Address List!")
            return
        else:
            ae = recipient.AddressEntry
            u = ae.GetExchangeUser()
            
        if 'EX' == ae.Type: email_address = u.PrimarySmtpAddress
        if 'SMTP' == ae.Type: email_address = ae.Address
        #https://social.msdn.microsoft.com/Forums/expression/en-US/9faa0862-4824-4691-8531-fe403a7eb3ff/how-can-i-addgo-back-to-150-office-library-references-after-2016-update?forum=accessdev
        #print(recipient.Name + ' is ' + 'sendable' if recipient.Sendable else 'not sendable')
        
        try:
            if BankID not in list(self.results.keys()):
                self.results[BankID] = {}
                self.results[BankID]['BankID'] = BankID
                self.results[BankID]['Name'] = recipient.Name 
                self.results[BankID]['Department'] = u.Department
                self.results[BankID]['JobTitle'] = u.JobTitle
                self.results[BankID]['CompanyName'] = u.CompanyName
                self.results[BankID]['City'] = u.City
                self.results[BankID]['StreetAddress'] = u.StreetAddress
                self.results[BankID]['OfficeLocation'] = u.OfficeLocation
                self.results[BankID]['MobileTelephoneNumber'] = u.MobileTelephoneNumber
                self.results[BankID]['email'] = email_address
                self.results[BankID]['LineManager'] = u.GetExchangeUserManager().Name
        except:
            pass
            
        if isManager:
            try:
                self.results[BankID]['workerBankID']
            except KeyError:
                self.results[BankID]['workerBankID'] = []
            
            self.results[BankID]['workerBankID'].append(workerBankID)
        
        for k,v in self.results[BankID].items():
            print(str(k) + ": " + str(v))
        print("==================================\n")

    def find(self, *args, **kwds):
        if(len(args)==1 and isinstance(args[0],str)):
            self._BankID = args[0]
            self._getfromOutlook(self._BankID)
        elif(len(args)==1 and isinstance(args[0],list)): 
            list_BankID = args[0]
            for BankID in list_BankID: self.find(BankID)
        else:
            print("BankID needs to be in a list or a str")
            
    def __find_manager(self, *args, **kwds):
        if(len(args)==1 and isinstance(args[0],str)):
            BankID = args[0]
        
        recipient = self._outlook.Session.CreateRecipient(BankID)
        rev = recipient.Resolve()
        if not rev: 
            print("BankID: " + str(BankID) + " is not found in Global Address List!")
            return
        else:
            ae = recipient.AddressEntry
            u = ae.GetExchangeUser()
            m = u.GetExchangeUserManager()
            # Manager's BankID
            try: 
                m_BankID = int(m.Address[ m.Address.rfind("/cn=")+4 : m.Address.rfind("/cn=")+4+7])
            except ValueError:
                print('Manager BankID is special in Address')
                m_BankID = m.Address[ m.Address.rfind("/cn=")+4 : len(m.Address)-3]
            except AttributeError:
                print('Invalid BankID!')
                return
                
            print("Manager's BankID: " + str(m_BankID) if not isinstance(m_BankID,str) else str(m_BankID) )
            self._getfromOutlook(str(m_BankID), isManager = True, workerBankID = str(BankID))
        
    def find_orgchart(self, *args, **kwds):
        if(len(args)==1 and isinstance(args[0],str)): 
            BankID = args[0]

        recipient = self._outlook.Session.CreateRecipient(BankID)
        rev = recipient.Resolve()
        if not rev: 
            print("BankID: " + BankID + " is not found in Global Address List!")
            return
        
        prevBankID = ''
        while BankID != prevBankID:
            u = self._outlook.Session.CreateRecipient(BankID).AddressEntry.GetExchangeUser()
            print(u.Name)
            prevBankID = BankID
            BankID = u.GetExchangeUserManager().Name

    def find_manager(self, *args, **kwds):
        if(len(args)==1 and isinstance(args[0],list)): 
            list_BankID = args[0]
            for BankID in list_BankID: self.__find_manager(BankID)
        elif(len(args)==1 and isinstance(args[0],str)): 
            list_BankID = args[0]
            self.__find_manager(list_BankID)
        else:
            print("BankID: " + BankID + " is not found in Global Address List!")
            
    def get_results(self):
        return pd.DataFrame(self.results).transpose()
    
    def del_results(self):
        self.results = {}
            
    __doc__ = """
    ======================================================
    USES
    ======================================================
    - Find contact details from Outlook Global Address List
    
    ======================================================
    SAMPLE
    ======================================================
    gal = outlookGAL()
    gal.find('1522918')
    gal.find_manager('1522918')
    gal.find('Winters, Bill')
    gal.find_manager('Winters, Bill')
    gal.find_manager('Vinals, Jose')
    gal.find_orgchart('1580873')
    gal.results
    
    gal = outlookGAL()
    list_BankID = ['1289066','1379266','1216415','1580873']
    gal.find_all(list_BankID)
    gal.results
    b=gal.get_results()
    """

#==============================================================================
# 
#==============================================================================

class FileWriter(object):
    
    def __init__(self, filename):
        self.file = open(filename, "w")

    def printlog(self, a_string):
        str_uni = a_string.encode('utf-8')
        self.file.write(str(str_uni))
        self.file.write("\n")
        print(a_string)

    def flush(self):
        self.file.flush()
    
    __doc__ = """
    ====================================================================
    USES
    ====================================================================
    - print string from python to log.txt saved on local drive
    ====================================================================    
    SAMPLE
    ====================================================================
    log = FileWriter("filename.txt")
    log.printlog("some string to log into .txt file")
    log.flush()
    """
        
        
class CheckMail(object):
    
    def __init__(self, email, mailbox, startDate=None, endDate=None, all_msgs = False): #startDate, endDate):
        self._log = FileWriter(re.sub("\.","",email[:email.find('@')]) + "_" + mailbox + ".txt")
        self._outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        self._account = self._outlook.Folders[email]
        self._mailbox = self._account.Folders[mailbox]
        self.results = [] # store log in dict {subject: body}
        self._startDate = startDate
        self._endDate = endDate
        self._all_msgs = all_msgs
        
        # check for dates
        try:
            self._startDate
            self._endDate
        except (NameError, UnboundLocalError) as e:
            print('Please enter in Start and End Date since you are not reading all messages!')

        if not isinstance(self._startDate, pd.Timestamp): self._startDate = pd.to_datetime(startDate)
        if not isinstance(self._endDate, pd.Timestamp): self._endDate = pd.to_datetime(endDate)
    
    
    def _logger(self, msg):
        #==============================================================================
        # Microsoft Outlook 16.0 Object Library Methods
        
        #mailbox="JustinSHuiMing.Yeo@sc.com"
        #folderindex = 'Inbox'
        #msgs=win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI").Folders[mailbox].Folders[folderindex].Items
        #msg = msgs.GetLast()

        #msg.Subject
        #msg.Body
        #msg.SenderName
        #msg.ReceivedTime
        
        #msg.Attachments.Count
        #for i in range(1, msg.Attachments.Count+1 ):
        #    msg.Attachments.Item(i).DisplayName
        #==============================================================================
        self._log.printlog("Message Received Time at: " + str(msg.ReceivedTime))
        self._log.printlog("Sender: " + msg.SenderName)
        self._log.printlog("Subject: " + msg.Subject)
        self._log.printlog("Body: " + msg.Body)
        
        print("Checking on... " + msg.Subject)
        
        atch = [] # attachment names
        if msg.Attachments.Count >=1 :
            for i in range(1,msg.Attachments.Count+1) :
                self._log.printlog("Attachment " + str(i) + " : " + msg.Attachments.Item(i).DisplayName)
                atch.append(msg.Attachments.Item(i).DisplayName)
        
        # save for return
        self.results.append({"Subject": msg.Subject, 
                             "Time": msg.ReceivedTime,
                             "Sender": msg.SenderName,
                             "Body": msg.Body,
                             "Attachments": atch
                             })
    
    def check(self):
        
        # initialise HEAD
        msgs = self._mailbox.Items
        try: 
            msg = msgs.GetLast() # most recent
        except: 
            print(mailbox + ' has no mails!')
            
        while msg:
            # limit to received DateTime to (startDate, endDate)
            if self._all_msgs == False: 
                # stop if beyond filter
                if msg.ReceivedTime.date() > self._endDate.date():
                    msg = msgs.GetPrevious()
                    continue
                if msg.ReceivedTime.date() < self._startDate.date(): 
                    break             
            
            self._logger(msg)
            msg = msgs.GetPrevious()
        
        self._log.printlog("-" * 50)
        self._log.flush()
        
        
    def find(self, pattern, where):
        if where=='Attachments': print("use .find_attachments for Attachments. Only for Body, Subject,... ")
        # find for pattern = specific keywords + variant from where = Subject, Body, ...             
        pattern = pattern.lower()
        hasit = [ True if pattern in m[where].lower() else False for m in self.results ]
        return hasit
    
    def get(self,what):
        return [x[what] for x in self.results]
    
#    def find_attachments(self, pattern, download=False):
#        att = self.get('Attachments')
#        cwd = os.getcwd()
#        hasit = [False for x in range(len(att))]
#        for j in range(len(att)):
#            if len(att[j])>=1:
#                for a in att[j]:
#                    if pattern.lower() in a.lower(): 
#                        hasit[j]
#                        # if need to download
#                        if download:
#                            cwd + "\\" + 
#                        continue
#        return hasit
    
    __doc__ = """
    ====================================================================
    USES
    ====================================================================
    - Find in subject, body, sendername whether it contains a string pattern
    ====================================================================    
    SAMPLE
    ====================================================================
    mail = CheckMail(email = email, mailbox = mailbox, all_msgs=True)
    mail = CheckMail(email = email, mailbox = mailbox, all_msgs=False, startDate = '20180105', endDate = '20180117')
    
    mail.check()
    
    mail.results
    len(mail.results)
    tf=mail.find(pattern = 'justin', where = 'Subject')
    list(map(mail.results.__getitem__, [i for i in range(len(tf)) if tf[i] is True] ))
    tf=mail.find(pattern = 'justin', where = 'Body')
    list(map(mail.results.__getitem__, [i for i in range(len(tf)) if tf[i] is True] ))
    for i in list(map(mail.results.__getitem__, [i for i in range(len(tf)) if tf[i] is True] )):
        print(i['Body'])
    """


#####################################
# AUTO_REPLY.PY
#####################################

import os
os.getcwd()
os.chdir('C:\\Users\\1580873\\Desktop\\RiskView')

import listen_python
import win32com.client
import pythoncom


#==============================================================================
# main script
#==============================================================================

if __name__ == "__main__":
    outlook_handler = listen_python.Handler_Class
    outlook_handler._IC = 'Justinshuiming.Yeo@sc.com'

    handler = win32com.client.DispatchWithEvents("Outlook.Application", outlook_handler)
    # http://docs.activestate.com/activepython/3.4/pywin32/pythoncom.html
    # win32com.client.DispatchWithEvents?
    pythoncom.PumpMessages()





#outlook = win32com.client.DispatchWithEvents("Outlook.Application", listen_python.Handler_Class)
## http://docs.activestate.com/activepython/3.4/pywin32/pythoncom.html
## win32com.client.DispatchWithEvents?
#pythoncom.PumpMessages()

#==============================================================================
# 
#==============================================================================

#class EventHandler:
#
#    def set_params(self, client):
#        self.client = client
#
#    def OnConnected(self):
#        print  "connected!"
#        self.client.do_something()
#        return True
#
#client = win32com.client.Dispatch("Lib.Obj")
#handler = win32com.client.WithEvents(client, EventHandler)
#handler.set_client(client)
#
#client.connect()
#
#while True:
#    PumpWaitingMessages()
#    time.sleep(1)








