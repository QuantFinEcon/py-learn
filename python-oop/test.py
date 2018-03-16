######################################

import pandas as pd
import numpy as np
import time
import string

#==============================================================================
# decorators
#==============================================================================
def timeit(method):
    def timed(*args, **kw):
        ts = time.time()
        result = method(*args, **kw)
        te = time.time()
        if 'log_time' in kw:
            name = kw.get('log_name', method.__name__.upper())
            kw['log_time'][name] = int((te - ts) * 1000)
        else:
            print( '%r  %2.2f ms' %  (method.__name__, (te - ts) * 1000))
        return result
    return timed

def dynamic_arg(config):
    def dynamic_arg_f(f):
        def f_decorator(*args, **kwargs):
            kwargs.update(config.__dict__)
            return f(*args, **kwargs)
        return f_decorator
    return dynamic_arg_f


class config_csv(object):
    def __init__(self, sep=",", header=0, skiprows=0, skipfooter=0, \
                 encoding='utf-8', dtype=None, names=None, \
                 usecols = None, low_memory=True):
        self.encoding = encoding #'iso-8859-1'
        self.sep = sep
        self.header = header
        self.skipfooter = skipfooter
        self.skiprows = skiprows
        self.dtype = dtype
        self.usecols = usecols
        self.names = names
        self.low_memory = low_memory
        
class config_excel(object):
    def __init__(self, sheet_name, header=0, skiprows=0, skip_footer = 0, \
                 encoding='iso-8859-1', dtype=None, names=None):
        self.encoding = encoding
        self.sheet_name = sheet_name
        self.skiprows = skiprows
        self.header = header
        self.skip_footer = skip_footer
        self.dtype = dtype
        self.names = names

                     
#==============================================================================
# classes
#==============================================================================

class pivottable(object):
    
    @timeit
    def __init__(self, filepath, dest_path, csv_config = None, excel_config = None):
        pd.set_option('display.float_format', lambda x: '%.5f' % x)
        self._writer = pd.ExcelWriter(dest_path) # read SFTP file
        self._temp = {}
        self._data = None
        
        @dynamic_arg(csv_config)
        def read_csv_with_config(filepath_or_buffer, **kwargs):
            return pd.read_csv(filepath_or_buffer=filepath_or_buffer, **kwargs)
        
        @dynamic_arg(excel_config)
        def read_excel_with_config(io, **kwargs):
            return pd.read_excel(io=io, **kwargs)

        if isinstance(filepath, list):
            # stitch multiple filepaths
            for files in filepath:
                print("merging " + str(files) )
                if files.split('.')[1] == 'csv':
                    if self._data is None:
                        self._data = read_csv_with_config(filepath_or_buffer = files)
                    else: 
                        # append
                        self._data = self._data.append(read_csv_with_config(filepath_or_buffer = files))
                elif files.split('.')[1] in ['xlsx','xls']:
                    if self._data is None:
                        self._data = read_excel_with_config(io=files)
                    else: 
                        # append
                        self._data = self._data.append(read_excel_with_config(io=files))

                else:
                    print('invalid filepath' + files + ' only reads data from csv/xlsx/xls filetypes')
                    return
                
                
                
        else:            
            if filepath.split('.')[1] == 'csv':
                self._data = read_csv_with_config(filepath_or_buffer = filepath)
            elif filepath.split('.')[1] in ['xlsx','xls']:
                self._data = read_excel_with_config(io=filepath)
            else:
                print('invalid filepath' + filepath + ' only reads data from csv/xlsx/xls filetypes')
                return
        
        # fix column names
        cv=[x.strip('.') for x in self._data.columns.values]
        cv=[x.replace(".."," ") for x in cv]
        cv=[x.replace("."," ") for x in cv]
        self._data = self._data.rename(columns = dict(zip( list(self._data.columns.values), cv)))
        print(list(self._data.columns.values))
    
    @timeit
    def sortPT(self, which):
        for k,v in self._temp.items():
            if which in v.columns.values:
                idx_col = list(v.index.names)
                self._temp[k] = v.reset_index().sort_values(which, ascending=False).set_index(idx_col)
                print(which + " column sorted from sheet: " + k)
            else:
                # do nothing
                print(which + " not in sheet: " + k)
    
    @timeit
    def create_table(self, reports):
        '''
        values is str, one column
        index can be str or list
        columns only for categorical columns... pivot out cats as columns
        aggfunc is function or list of functions
        '''
        for k,v in reports.items():
            if k.startswith('#'): continue # not pivot fields...
            print("Working on " + k + " pivot table...")
            
            # get data and filter
            f=v['filter']
            if len(f)>=1:
                fc=''
                for f1,f2 in f.items():
                    fc += 'self._data[' +"'"+ str(f1) +"'"+ '].isin(' + str(f2) + ')'
                    fc += ' & '
                fc = fc.strip(' & ')
                filtered_data = eval('self._data.loc['+ fc +']')
            else:
                filtered_data = self._data
            
            #pivot
            #store in self._temp dictionary before writing into excel
            pt = pd.pivot_table(filtered_data, values=v['val'], index=v['idx'], \
                      columns=v['col'], aggfunc= v['aggfunc'],\
                      margins=True,dropna=True,margins_name='Total',
                      fill_value=0)
            
            if isinstance(v['aggfunc'], list):
                pt.columns = [x[1] + " " + x[0] for x in pt.columns.values]
                
            self._temp[k] = pt
            
            pass
        
    @timeit
    def drop_col(self, which):
        for k,v in self._temp.items():
            if which in v.columns.values:
                v.__delitem__(which)
                print(which + " column drop from sheet: " + k)
            else:
                # do nothing
                print(which + " not in sheet: " + k)
        
    @timeit
    def drop_row(self, which):
        for k,v in self._temp.items():
            if which in v.columns.values:
                v = v.drop(which)
                print(which + " row drop from sheet: " + k)
            else:
                # do nothing
                print(which + " not in sheet: " + k)    
    
    @timeit
    def diff(self, post, pre):
        for k,v in self._temp.items():
            if post and pre in v.columns.values:
                v[post+" - "+pre] = v[post] - v[pre]
                v[v == np.Inf] = np.NaN
                print(post+" - "+pre + " done in sheet: " + k)
            else:
                # do nothing
                print(pre + " and " + post + " not in sheet: " + k)

    @timeit
    def diff_abs(self, post, pre):
        for k,v in self._temp.items():
            if post and pre in v.columns.values:
                v[post+" - "+pre+" (abs)"] = np.abs( v[post] - v[pre] )
                v[v == np.Inf] = np.NaN
                print(post+" - "+pre+" (abs)" + " done in sheet: " + k)
            else:
                # do nothing
                print(pre + " and " + post + " not in sheet: " + k)

    @timeit
    def diff_perc(self, post, pre):
        for k,v in self._temp.items():
            if post and pre in v.columns.values:
                v[post+" - "+pre+" (%)"] = \
                (v[post] - v[pre]).to_frame().divide(v[pre], axis = 'rows')
                v[v == np.Inf] = np.NaN
                print(post+" - "+pre+" (%)" + " done in sheet: " + k)
            else:
                # do nothing
                print(pre + " and " + post + " not in sheet: " + k)
                
    @timeit
    def normalise(self, which, by='columns'):
        # normalise to a base, either row/s or column/s,  0-100%
        # if by=columns, which can be a str column name or a list of str column names
        for k,v in self._temp.items():
            
            if by.startswith('c'):
                if isinstance(which,str):
                    total = v.loc['Total',:][which][0]
                    v[which + ' contribution to column total'] = \
                    v.loc[:,which].to_frame().divide(total)
                    v[v == np.Inf] = np.NaN
                    print('[' + str(which) + " contribution to column total] done in sheet: " + k)
                elif isinstance(which,list):
                    total = v.loc['Total',:][which].values.tolist()[0]
                    v[ [w + ' contribution to column total' for w in which] ] = \
                    v.loc[:,which].divide(total)
                    v[v == np.Inf] = np.NaN
                    print('[' + str(which) + " contribution to column total] done in sheet: " + k)
                else:
                    print('[' + str(which) + " contribution to column total] NOT done in sheet: " + k)

            elif by.startswith('r'):
                if isinstance(which,str):
                    v[which + ' contribution to row total'] = \
                    v.loc[:,which].to_frame().divide(v.loc[:,'Total'], axis='rows')
                    v[v == np.Inf] = np.NaN
                    print('[' + str(which) + " contribution to row total] done in sheet: " + k)
                elif isinstance(which,list):
                    v[ [w + ' contribution to row total' for w in which] ] = \
                    v.loc[:,which].divide(v.loc[:,'Total'], axis='rows')
                    v[v == np.Inf] = np.NaN
                    print('[' + str(which) + " contribution to row total] done in sheet: " + k)
                else:
                    print('[' + str(which) + " contribution to row total] NOT done in sheet: " + k)

            else:
                print('Only can normalise by row or columns. by = row or columns')    
    
    def results(self):
        return self._temp
    
    @timeit
    def closewb(self):
        # write pivot table into a sheet
        # https://pandas.pydata.org/pandas-docs/stable/options.html
        for k,v in self._temp.items():
            print('writing sheet '+ k + '...')
            v.to_excel(self._writer, sheet_name = k, 
                       index = True, header = True, na_rep='', inf_rep='',
                       merge_cells = False, float_format = '%.2f',
                       encoding = 'iso-8859-1')
        self._writer.close()


    __doc__ = """
    ==========================================
    USES
    ==========================================
    - to generate pivot tables with a set of pivot table fields
      such as columns, indexes, filters, values, aggFunc from a JSON format .txt file
    - reads data from csv/excel and write into a new excel file
      each with a sheet representing the pivot table description
    ==========================================
    SAMPLE CODE
    ==========================================
    # pivot fields as dictionary
    reports  = \
   {'cpty': {'val': 'Exposure_for_RWA',
       'idx': ['Country_of_Exposure','Counter Party','Standard Product'],
       'col': ['Reporting month'],
       'filter': {'STDF_Template_Indicator': ['STDF-RiskMeasures']},
       'aggfunc': np.sum
       },
    'total': {'val': 'Outstanding',
       'idx': ['Orgnisational_unit_level_1'],
       'col': ['Reporting month'],
       'filter': {'STDF_Template_Indicator': ['STDF-RiskMeasures'],
                  'BS_TYPE':['ASS','DAF','DAO','DLO','DLF']},
       'aggfunc': np.sum
       },
    'product_group': {'val': 'Exposure_for_RWA',
       'idx': ['Orgnisational_unit_level_1','Mgt Prod Grouping'],
       'col': ['Reporting month'],
       'filter': {'STDF_Template_Indicator': ['STDF-RiskMeasures']},
       'aggfunc': np.sum
       },
    'product': {'val': 'Exposure_for_RWA',
       'idx': ['Orgnisational_unit_level_1','Standard Product'],
       'col': ['Reporting month'],
       'filter': {'STDF_Template_Indicator': ['STDF-RiskMeasures']},
       'aggfunc': np.sum
       },
    'client_group': {'val': 'Exposure_for_RWA',
       'idx': ['Client Group'],
       'col': ['Reporting month'],
       'filter': {'STDF_Template_Indicator': ['STDF-RiskMeasures']},
       'aggfunc': np.sum
       },
    'credit_grade': {'val': 'Exposure_for_RWA',
       'idx': ['CG Final'],
       'col': ['Reporting month'],
       'filter': {'STDF_Template_Indicator': ['STDF-RiskMeasures']},
       'aggfunc': np.sum
       }
    }

    p=SLA_control_pivot(filepath = 'C:\\Users\\1580873\\Desktop\\RECON 23 Jan 2018\\recon.csv', 
                    dest_path = 'C:\\Users\\1580873\\Desktop\\RECON 23 Jan 2018\\report.xlsx')
        
    p.create_table(reports)

    p.diff_abs(pre='17-Jun', post='17-Sep')
    p.diff_perc(pre='17-Jun', post='17-Sep')

    p.normalise(which=['17-Jun','17-Sep'], by='columns')
    p.normalise(which=['17-Jun','17-Sep'], by='rows')

    p.drop_col('Total')
    p.drop_row('Total')
    
    p.sortPT('17-Sep - 17-Jun (abs)')

    #commit calculations to excel
    temp=p.results()
    p.closewb()
    """
    
class excel_editor(object):

    @timeit
    def __init__(self, filepath, visibility=False):
        import win32com.client
        self._excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
        self._excel.Visible = visibility
        self._wb = self._excel.Workbooks.Open(filepath)
        self.sheetcount = self._wb.Worksheets.Count
        self.sheetnames = [self._wb.Worksheets(i).Name for i in range(1,self.sheetcount+1)]
    
    @timeit    
    def scan_columns(self, header=1):
        for i in range(1,self.sheetcount+1):
            ws = self._wb.Worksheets(i)
            print(ws.Name)
            print([x for x in list(ws.Rows( str(header)+':'+str(header) ).Value[0]) if x is not None])
            print("\n")
            
    @timeit    
    def scan_dtypes(self, header=1):
        for i in range(1,self.sheetcount+1):
            ws = self._wb.Worksheets(i)
            print(ws.Name)
            print([x for x in list(ws.Rows( str(header)+':'+str(header) ).Value[0]) if x is not None])
            print("\n")
    
    @timeit    
    def autofit_columns(self):
        for i in range(1,self.sheetcount+1):
            ws = self._wb.Worksheets(i)
            ws.Columns.AutoFit()

    @timeit            
    def format_cells(self, which, style, decimal_places=2, header=1):
        # which is a list of column names
        # style is Comma Percent Dollar
        if isinstance(which, str): which = list(which)
        for i in range(1,self.sheetcount+1):
            ws = self._wb.Worksheets(i)
            for w in which:
                try:
                    # Default exact match
                    where = ws.Rows(str(header)+':'+str(header)).Find(What=w).Column
                    let=list(string.ascii_uppercase)[where-1]
                    print(let+':'+let)
                    ws.Columns(let+':'+let).Style = style
                    if style == 'Percent':
                        ws.Columns(let+':'+let).NumberFormat  = '0.' + '0'*decimal_places + '%'              
                except:
                    print('No such column ' + w)
                        
    @timeit
    def close(self):
        try:
            self._wb.Save()
            self._excel.Application.Quit()
        except:
            print('No excel file to close!')

    __doc__ = """
    ==========================================
    USES
    ==========================================
    - format excel ranges e.g. cell style, autofit columns width
    ==========================================
    SAMPLE CODE
    ==========================================
    e = pivot_table.excel_editor(filepath, visibility = False)
    e.autofit_columns()
    e.scan_columns()
    e.format_cells(which=['17-Sep','17-Jun','17-Sep - 17-Jun (abs)','17-Sep - 17-Jun'], style="Comma")
    e.format_cells(which=['17-Sep - 17-Jun (%)','17-Sep - 17-Jun (abs) contribution to column total'], style="Percent", decimal_places=5)
    e.close()
    """


#==============================================================================
# parsing inputs from JSON
#==============================================================================
    
# read all python functions and objects stored as value in json keys
def recursive_items(dictionary):
    for key, value in dictionary.items():
        if isinstance(value, dict): 
            recursive_items(value)
        else: 
            print("key = " + str(key))
            print("value = " + str(value) )
            print("value dtype = " + str(type(value)) )            
            try:
                print(eval( str(value) ))
            except NameError:
                print("Cannot eval str")


#https://stackoverflow.com/questions/12507206/python-recommended-way-to-walk-complex-dictionary-structures-imported-from-json
def walk_for_functions(d, path = [], paths_to_chg = {}):
    '''
    traverse all paths in dict and returns all paths 
    that need its values to be reassigned
    to evaluated string from JSON
    ## NOTES ##
     - for list in JSON, its ["abc"] not "[abc]" since load.JSON can read list []
     - single quotes inside functions "post":"Template("$post")"
    '''
    for k,v in d.items():
        if isinstance(v, str):
            try:
                v1 = eval(v)
                if not isinstance(v1,int):
                    if str(v1).startswith('<') and str(v1).endswith('>'):
                        path.append(k)
                        print("{} <--- {}".format("|".join( [str(x) for x in path] ), str(v)))
                        
                        nesting = ''
                        for k in path: nesting += "['" + k + "']"
                        paths_to_chg[nesting] = v
                        path.pop()
                    
            except NameError:
                pass
            
            except SyntaxError:
                pass
        
        elif isinstance(v, list):
            for i in range(len(v)):
                try:
                    v1 = eval(v[i])
                    if not isinstance(v1,int):
                        if str(v1).startswith('<') and str(v1).endswith('>'):
                            path.append( i )
                            print("{} <--- {}".format("|".join( [str(x) for x in path] ), str(v[i])))
                            
                            nesting = ''
                            for k in path: 
                                if isinstance(k, int): nesting += "[" + str(k) + "]"
                                else: nesting += "['" + k + "']"
                            paths_to_chg[nesting] = v[i]
                            path.pop()
                        
                except NameError:
                    pass
                
                except SyntaxError:
                    pass

        
        elif v == {}:
            pass
            #path.append(k)
            #path.pop()
            
        elif isinstance(v, dict):
            path.append(k)
            walk_for_functions(v)
            path.pop()
            
        elif isinstance(v, int):
            pass
        
        else:
            print("###Type {} not recognized: {}.{}={}".format(type(v), "|".join(path),k, v))

    return paths_to_chg



def walk_for_obj(d, obj, kw, path = [], paths_to_chg = {}):
    '''
    traverse all paths in dict and returns all paths 
    that need its values to be reassigned
    to evaluated string from JSON
    ## NOTES ##
     - for list in JSON, its ["abc"] not "[abc]" since load.JSON can read list []
     - single quotes inside functions "post":"Template("$post")"
    '''
    for k,v in d.items():
        if isinstance(v, obj):
            
            keys = list(set([ x[1] for x in Template.pattern.findall(v.template)]))
            arg = ''
            print(keys)
            for key in keys:
                arg += key + "=kw['" + key + "'], "
            if len(arg) != 0: arg = arg.rstrip(', ')
            
            v = eval("v.substitute(" + arg + ")")
            
            path.append(k)
            print("{} <--- {}".format("|".join( [str(x) for x in path] ), str(v)))
            
            nesting = ''
            for k in path: nesting += "['" + k + "']"
            paths_to_chg[nesting] = v
            path.pop()
            

        elif isinstance(v, list):
            for i in range(len(v)):
                if isinstance(v[i], obj):
                    
                    keys = list(set([ x[1] for x in Template.pattern.findall(v[i].template)]))
                    arg = ''
                    for key in keys:
                        arg += key + "=kw['" + key + "'], "
                    if len(arg) != 0: arg = arg.rstrip(', ')
                    
                    v[i] = eval("v[i].substitute(" + arg + ")")
                    
                    path.append(k)
                    print("{} <--- {}".format("|".join( [str(x) for x in path] ), str(v[i])))
                    
                    nesting = ''
                    for k in path: nesting += "['" + k + "']"
                    paths_to_chg[nesting] = v[i]
                    path.pop()
        
        elif v == {}:
            pass
            #path.append(k)
            #path.pop()
            
        elif isinstance(v, dict):
            path.append(k)
            walk_for_obj(d=v, kw=kw, obj=obj)
            path.pop()

        elif isinstance(v, str):
            pass
                    
        elif isinstance(v, int):
            pass
        
        else:
            print("###Type {} not recognized: {}.{}={}".format(type(v), "|".join(path),k, v))

    return paths_to_chg


def parse_templates(dictionary, keywords):
    paths_to_chg = walk_for_obj(d=dictionary, kw=keywords, obj=string.Template)
    
    for k,v in paths_to_chg.items():
        #print(k,v)
        print( "dictionary" + k + ' = "' + v + '"')
        exec("dictionary" + k + ' = "' + v + '"',None,locals())

    return dictionary

# https://stackoverflow.com/questions/2220699/whats-the-difference-between-eval-exec-and-compile-in-python
# https://www.saltycrane.com/blog/2008/01/python-variable-scope-notes/
def eval_dict_values(dictionary):
    
    paths_to_chg = walk_for_functions(dictionary)
    
    for k,v in paths_to_chg.items():
        #print(k,v)
        print( "dictionary" + k + ' = eval("' + v + '")')
        exec("dictionary" + k + ' = eval("' + v + '")',None,locals())

    return dictionary



##==============================================================================
## namepsace
##==============================================================================
#def function(v):
#    _locals = locals()
#    print(_locals)
#    _locals['x'] = 1
#    abc=100 in _locals
#    df = 50
#    print(_locals)
#    exec("x = v", None, _locals)
#    print(locals())
#    return _locals
#
#d = {'a':1,'b':2}
#
#function(2)
#
#
#value='np.sum'
#eval(value)
#exec("value=eval(value)")
#value
#
#
#eval("5")
#eval("abc")
#eval("np.sum")
#eval('Template("$Name")')
#str(eval('Template("$Name")')) # starts with <
#
#t=eval('Template("$Name is an idiot")')
#t.substitute(Name='Justin')
#
#value = 'np.sum'
#exec("report['RWA_product_group']['aggfunc'] = exec(value)") # this dont work
#exec("report['RWA_product_group']['aggfunc'] = eval(value)")


#####################

import os
import pandas as pd
import numpy as np
import difflib
import itertools
import time


#==============================================================================
# MANUAL TESTING
#==============================================================================

os.getcwd()
os.chdir('C:\\Users\\1580873\\Desktop\\Completed Developments\\Utility\\Source')

#sep=
#encoding=

new_filepath = 'sample_new.csv'
old_filepath = 'sample_old.csv'

new = pd.read_csv(new_filepath , sep=',', encoding='utf-8')
old = pd.read_csv(old_filepath , sep=',', encoding='utf-8')

print(str(os.stat(new_filepath).st_size/1000000) +"MB")#mb
print(str(os.stat(old_filepath).st_size/1000000) +"MB")#mb

comparator = compare_table(new=new, old=old)

comparator.shape_diff()
comparator.find_diffcol()
comparator.get_common_cols()

comparator.find_dupPK(table=new, PK='Main.Profile.Code')
out=comparator.find_colmembership()
comparator.dict_to_xls(out, filename = "colmember.xlsx")

#mutators
comparator.set_common_cols()
comparator.remove_duplicates()

comparator.values_change()
comparator.values_change_with_matching()
# static: for any other tables 
comparator.from_to(left=new, right=old) 

comparator.compare_innerjoin_exclusions()





#==============================================================================
# HELPERS
#==============================================================================

def replicate_firstcopy_dtype(a, replace='object', to='str'):
    # mixed type NaN with float
    dtypes = a.dtypes
    k=[ str(x) for x in dtypes.index ]
    v=[ str(x) for x in dtypes.values ]
    list(map(lambda x: x=='object', v))
    v2 = [to if x=='object' else x for x in v]
    return dict(zip(k,v2))


def timeit(method):
    def timed(*args, **kw):
        ts = time.time()
        result = method(*args, **kw)
        te = time.time()
        if 'log_time' in kw:
            name = kw.get('log_name', method.__name__.upper())
            kw['log_time'][name] = int((te - ts) * 1000)
        else:
            print( '%r  %2.2f ms' %  (method.__name__, (te - ts) * 1000))
        return result
    return timed

class compare_table(object):

    @timeit
    def __init__(self, new, old):
        self.t1 = new # NEW #pandas.DataFrame # LEFT 
        self.t2 = old # OLD # RIGHT
        # pre comparison cleaning
        
    @timeit
    def get_common_cols(self):
        s1=list(self.t1.columns.values)
        s2=list(self.t2.columns.values)
        return [x for x in list(self.t1.columns.values) if x in list(set(s1) & set(s2))]
    
    @timeit        
    def set_common_cols(self):
        s1=list(self.t1.columns.values)
        s2=list(self.t2.columns.values)
        self.t1 = self.t1[[x for x in list(self.t1.columns.values) if x in list(set(s1) & set(s2))]]
        self.t2 = self.t2[[x for x in list(self.t2.columns.values) if x in list(set(s1) & set(s2))]]
        pass
    
    # troubleshoot why PK is not unique?
    @staticmethod
    def find_dupPK(table, PK=None):        
        df = pd.DataFrame(columns=table.columns.values)
        for d in table[PK][ table[PK].duplicated() ].unique():
            df = pd.concat([df, table.loc[table[PK]==d, :]],axis=0)
        return df
        
    @timeit            
    def remove_duplicates(self):        
        old_rows = self.t1.shape[0]
        self.t1 = self.t1.drop_duplicates(keep='first')
        print("Removed " + str(old_rows - self.t1.shape[0]) + " rows from table1 (New)")
        
        old_rows = self.t2.shape[0]
        self.t2 = self.t2.drop_duplicates(keep='first')
        print("Removed " + str(old_rows - self.t2.shape[0]) + " rows from table1 (Old)")
        pass
        
    # versioning control: test change of values for an EXACT similar table    
    @staticmethod 
    def from_to(left, right):
        out = {}
        left = left.reset_index()
        right = right.reset_index()
        boo = left.ne(right)
        idx = np.where(boo) # (row, col)
        
        try:
            # MUST BE SAME SHAPE
            chg_from = right.values[idx] # RIGHT is OLD. so from
            chg_to = left.values[idx]
        except IndexError:
            print('Ensure 2 tables in comparison are of the same shapes')
            return
        
        changed = boo.stack()[boo.stack()]
        changed.index.names = ['index','columns']
        # False positive found e.g. of 100 columns, 1 col is different --> 0.99>0.95 similarity but its a valid difference
        # how to filter out false positive (error)?
        
        # pass in "similarity" ratio to output for subjectivity evaluation
        # output the rows with mismatches too
        out['from_to_idx'] = pd.DataFrame({'from': chg_from, 'to': chg_to}, index=changed.index).dropna(axis=0, how='all')
        out['suspected_from'] = right.iloc[list(set([x[0] for x in out['from_to_idx'].index.values])),:]
        out['suspected_to'] = left.iloc[list(set([x[0] for x in out['from_to_idx'].index.values])),:]
        return out
        
    # but order or rows and columns must be the same    
    @timeit        
    def values_change_with_matching(self):
        '''
        sort rows by key(s)
        sort columns (only for common columns)
        '''
        idx = self.isolate_similar_rows(self, left=self.t1, right=self.t2, threshold=0.95)
        cc = self.get_common_cols()
        return self.from_to(left=self.t1[cc].iloc[idx['left'],:], \
                       right=self.t2[cc].iloc[idx['right'],:])
            
    @timeit        
    def values_change(self):        
        return self.from_to(self.t1, self.t2)
    
    # alternative use: from_to between 2 tables from left and right not from inner join    
    @timeit        
    def compare_innerjoin_exclusions(self, on=None):
        
        excl = self.innerjoin_exclusions(on)
        left = excl['left_only'] #ONLY IN NEW
        right = excl['right_only'] #ONLY IN OLD
        idx = self.isolate_similar_rows(self, left,right,threshold=0.95)
        # t1.iloc[[0,0,0],:] allows one-many
        return self.from_to(left=left.iloc[idx['left'],:], right=right.iloc[idx['right'],:])
    
    #which columns values is a miss? What is a 'inexact 'match?
    @staticmethod
    def similarity(row1, row2):
        '''
        useful if there are many redundant columns values
        0 to 1.00
        threshold: 0.95
        '''
        a="".join(str(x) for x in row1)
        b="".join(str(x) for x in row2)
        return difflib.SequenceMatcher(None, a, b).ratio()
    
    # DEPENDS on no. of columns and how much FP alpha tolerated
    @staticmethod
    def isolate_similar_rows(self,left,right,threshold=0.95):
        # NOTE!! Only use for subset of data.. otherwise, permutation will have Inf possibilties
        print('Please Wait. Script is running to match similar rows...')
        print('Definition of "Similar" Threshold is ' + str(threshold) + '/1.00')
        idx = pd.DataFrame([ [ x[0],x[1] ] for x in itertools.product(list(range(left.shape[0])), \
                            list(range(right.shape[0]))) ], \
                            columns = ['left','right'])
        idx['similarity'] = None
        
        for i in range(idx.shape[0]):
            li = idx.loc[i,'left']; ri = idx.loc[i,'right']
            idx.loc[i,'similarity'] = self.similarity(left.iloc[li,:],right.iloc[ri,:])
        
        return idx.loc[idx['similarity']>=threshold]
        

    #test row counts
    @timeit
    def shape_diff(self):
        r1=self.t1.shape[0]; c1=self.t1.shape[1]
        r2=self.t2.shape[0]; c2=self.t2.shape[1]
        print('t1 shape is ' + str(self.t1.shape) )
        print('t2 shape is ' + str(self.t2.shape) )
        print('t1 has ' + str(r1-r2) + ' more rows than t2')
        print('t1 has ' + str(c1-c2) + ' more columns than t2')
        pass
    
    #find columns not in both tables. Whats in A not in B? Whats in B not in A?
    @timeit
    def find_diffcol(self):
        c1=set(self.t1.columns.values)
        c2=set(self.t2.columns.values)
        
        notin1 = c1.difference(c2)
        notin2 = c2.difference(c1)
        
        if notin1 != set(): 
            print(notin1 + " not in table1 columns but in table2 columns")
        if notin2 != set(): 
            print(notin2 + " not in table2 columns but in table1 columns")
        if notin1 == set() and notin2 == set():
            print('No Differences in Columns!')
        pass
    
    #by column, test matchness between 1-1 row for those not not inner join... 
    @timeit
    def find_colmembership(self):
        '''
        returns a dictonary
        k1 common members per column
        k2 only in new, not in old
        k3 only in old, not in new
        '''
        cc = self.get_common_cols()
        
        print(self.t1[cc].dtypes)
        # by float / int not always True.. so go by length of unique values
        #str(t1['Exposure_for_RWA'].dtype).startswith('float')
        # how to handle float accuracy ???
        
        dic = {}
        for c in cc:
            u = list(self.t1[c].unique())
            dic[c] = u        
        # pretify into df
        u_t1_df = pd.DataFrame( dict([ (k,pd.Series(v)) for k,v in dic.items() ]) ).replace(np.nan,'',regex=True)
        
        dic = {}
        for c in cc:
            u = list(self.t2[c].unique())
            dic[c] = u        
        # pretify into df
        u_t2_df = pd.DataFrame( dict([ (k,pd.Series(v)) for k,v in dic.items() ]) ).replace(np.nan,'',regex=True)

        out = {}
        # k1
        dic = {}
        for c in cc:
            u=list(set(u_t1_df[c].unique()) & set(u_t2_df[c].unique()))
            if '' in u: u.remove('')
            dic[c] = u
        out['common'] = pd.DataFrame( dict([ (k,pd.Series(v)) for k,v in dic.items() ]) ).replace(np.nan,'',regex=True)

        # k2
        dic = {}
        for c in cc:
            u=list(set(u_t1_df[c].unique()) - set(u_t2_df[c].unique()))
            if '' in u: u.remove('')
            dic[c] = u
        out['newonly'] = pd.DataFrame( dict([ (k,pd.Series(v)) for k,v in dic.items() ]) ).replace(np.nan,'',regex=True)
        
        # k3
        dic = {}
        for c in cc:
            u=list(set(u_t2_df[c].unique()) - set(u_t1_df[c].unique()))
            if '' in u: u.remove('')
            dic[c] = u
        out['oldonly'] = pd.DataFrame( dict([ (k,pd.Series(v)) for k,v in dic.items() ]) ).replace(np.nan,'',regex=True)

        return out
    
    @timeit
    def innerjoin_exclusions(self, on=None):
        return self._compare_exclusions(self.t1, self.t2, on=on)

    #find rows not in inner join. Whats in A not in B? Whats in B not in A?
    def _compare_exclusions(self,t1, t2, on=None):
        '''
        returns a dictonary
        k1 in new, not in old - LEFT JOIN exclude INNER JOIN
        k2 in old, not in new - RIGHT JOIN exclude INNER JOIN
        '''
        out = {}
        XandY_diff = True # duplicated column names ending with _x LEFT and _y RIGHT
        if on==None: 
            on = self.get_common_cols()
            XandY_diff = False

        #check is ON is a unique KEY
        if any(t1.drop_duplicates()[on].duplicated()) or any(t2.drop_duplicates()[on].duplicated()):
            print("primary (or composite) Key is NOT UNIQUE! Try another set of \
                  columns as Key or use all (non numerical) columns.")
            return
        # self.remove_duplicates()
        t1 = t1.drop_duplicates(keep='first')
        t2 = t2.drop_duplicates(keep='first')
                
        # _merge both left_only right_only
        # in new not in old
        left = pd.merge(left=t1, right=t2, on=on, how='left', indicator=True)
        left_col = [True if i.endswith('_x') else False for i in list(left.columns.values)] \
            if XandY_diff else [True for i in list(left.columns.values)] 
        left = left.loc[left['_merge']=='left_only', left_col[:-1]]
        out['left_only'] = left

        # in old not in new
        right = pd.merge(left=t2, right=t1, on=on, how='left', indicator=True)
        right_col = [True if i.endswith('_y') else False for i in list(right.columns.values)] \
            if XandY_diff else [True for i in list(right.columns.values)] 
        right = right.loc[right['_merge']=='left_only', right_col[:-1]]
        out['right_only'] = right

        return out
    
    
    @staticmethod
    def dict_to_xls(d, filename=None):

        if filename == None: filename = 'exclusions.xlsx'
        writer = pd.ExcelWriter(path=os.getcwd()+'\\'+filename)
        
        for k,v in d.items(): 
            #if 'Index' in v.columns.values: v = v.drop('Index',axis=1)
            v.to_excel(writer, sheet_name = k, 
                           index = True, header = True, na_rep='', inf_rep='',
                           merge_cells = False, float_format = '%.2f',
                           encoding = 'iso-8859-1')
        print('Written into ' + filename)
        writer.close()
        pass
    
    
    __doc__ = \
    """
    ============================
    EXAMPLES
    ============================
    comparator = compare_table(new=new, old=old)
    
    comparator.shape_diff()
    comparator.find_diffcol()
    comparator.get_common_cols()
    
    comparator.find_dupPK(table=new, PK='Main.Profile.Code')
    out=comparator.find_colmembership()
    comparator.dict_to_xls(out, filename = "colmember.xlsx")
    
    #mutators
    comparator.set_common_cols()
    comparator.remove_duplicates()
    
    comparator.values_change()
    comparator.values_change_with_matching()
    # static: for any other tables 
    comparator.from_to(left=new, right=old) 
    
    comparator.compare_innerjoin_exclusions()

    
    ============================
    THOUGHTS ON RECONCILIATION
    ============================
    pre comparison cleaning:
        - find commmon columns
        - order columns
        - trim ends, cleanup column names
        - does columns data types match? standardise dtypes
        - drop indexes like S/N 1,2,3..., no purpose other than count

    comparison exceptions handling:
        - treatment of nan, nan == nan is False pd.fillna('')
        - trim all spaces at ends
        - exactness of numbers decimals? tolerance at what dp?
        - columns present but missing values at other table...
        - can comparison be done by ignoring columns with missing values? 
        - need one column match AT LEAST!
        
    """





############################################3


#==============================================================================
# JOB: CREATE generic pivot tables with filters, columns, indexes, values, aggregation fn
# last edited: 14 feb 2018, 1580873
# Inputs: key-value py dict.txt input path|src path|dest path
# C:/MFU/Manual Source|022018|C:/MFU/Manual
#==============================================================================

import os
import sys
import pandas as pd
import numpy as np
import time
import string
from string import Template
import json
import re
from copy import deepcopy
from datetime import datetime

#inputs
#print(sys.argv)
#pylibpath = sys.argv[1] # for importing
#config = sys.argv[2] # k-v .json path
#src = sys.argv[3] # src path
#dest = sys.argv[4] # dest path
#inputs = sys.argv[5] # for JSON Template

# manual input for testing
path = 'C:\\Users\\1580873\\Desktop\\Completed Developments\\Utility'
pylibpath  = path+'\\PyScript\\'
config = path+'\\Inputs\\MFU_DQC.json'
src = [path + '\\Source\\ALL_CRISKREPMFU_OtherBBExposure_20180223_D_001.csv', \
       path + '\\Source\\ALL_CRISKREPMFU_OtherBBExposure_20180226_D_001.csv']
dest = path + '\\Data\\mfu.xlsx'
inputs = "pre=2018-02-23; post=2018-02-26; CoB=20170226"

inputs = [re.sub(r'\s+', '', x) for x in inputs.split(sep=';')]
keywords = {}

for i in inputs:
    kv = i.split('=')
    keywords[kv[0]] = kv[1]

# CoB is yyyymmdd
cob = datetime.strptime(keywords['CoB'], "%Y%m%d")
if 'CoB' in keywords.keys():
    keywords['yy'] = datetime.strftime(cob, "%y")
    keywords['yyyy'] = datetime.strftime(cob, "%Y")
    keywords['mm'] = datetime.strftime(cob, "%m")
    keywords['mmm'] = datetime.strftime(cob, "%b")
    keywords['dd'] = datetime.strftime(cob, "%d")
    

os.chdir(pylibpath)
import pivot_table 
#from importlib import reload
#reload(pivot_table)

#==============================================================================
# 
#==============================================================================

#https://stackoverflow.com/questions/12507206/python-recommended-way-to-walk-complex-dictionary-structures-imported-from-json
def walk_for_functions(d, path = [], paths_to_chg = {}):
    '''
    traverse all paths in dict and returns all paths 
    that need its values to be reassigned
    to evaluated string from JSON
    ## NOTES ##
     - for list in JSON, its ["abc"] not "[abc]" since load.JSON can read list []
     - single quotes inside functions "post":"Template("$post")"
    '''
    for k,v in d.items():
        if isinstance(v, str):
            try:
                v1 = eval(v)
                if not isinstance(v1,int):
                    if str(v1).startswith('<') and str(v1).endswith('>'):
                        path.append(k)
                        print("{} <--- {}".format("|".join( [str(x) for x in path] ), str(v)))
                        
                        nesting = ''
                        for k in path: nesting += "['" + k + "']"
                        paths_to_chg[nesting] = v
                        path.pop()
                    
                    #for "None" --> --> None
                    if v1 is None:
                        path.append(k)
                        print("{} <--- {}".format("|".join( [str(x) for x in path] ), str(v)))
                        
                        nesting = ''
                        for k in path: nesting += "['" + k + "']"
                        paths_to_chg[nesting] = None
                        path.pop()
                    
            except NameError:
                pass
            
            except SyntaxError:
                pass
        
        elif isinstance(v, list):
            for i in range(len(v)):
                try: 
                    if not isinstance(v[i],int):
                        v1 = eval(v[i])
                        if not isinstance(v1,int):
                            if str(v1).startswith('<') and str(v1).endswith('>'):
                                path.append( k )
                                path.append( i )
                                print("{} <--- {}".format("|".join( [str(x) for x in path] ), str(v[i])))
                                
                                nesting = ''
                                for p in path: 
                                    if isinstance(p, int): nesting += "[" + str(p) + "]"
                                    else: nesting += "['" + p + "']"
                                paths_to_chg[nesting] = v[i]
                                path.pop()
                                path.pop()
                        
                except NameError:
                    pass
                
                except SyntaxError:
                    pass

        
        elif v == {}:
            pass
            #path.append(k)
            #path.pop()
            
        elif isinstance(v, dict):
            path.append(k)
            walk_for_functions(v)
            path.pop()
            
        elif isinstance(v, int):
            pass
        
        else:
            print("###Type {} not recognized: {}.{}={}".format(type(v), "|".join(path),k, v))

    return paths_to_chg



def walk_for_obj(d, obj, kw, path = [], paths_to_chg = {}):
    '''
    traverse all paths in dict and returns all paths 
    that need its values to be reassigned
    to evaluated string from JSON
    ## NOTES ##
     - for list in JSON, its ["abc"] not "[abc]" since load.JSON can read list []
     - single quotes inside functions "post":"Template("$post")"
    '''
    for k,v in d.items():
        if isinstance(v, obj):
            
            keys = list(set([ x[1] for x in Template.pattern.findall(v.template)]))
            arg = ''
            print(keys)
            for key in keys:
                arg += key + "=kw['" + key + "'], "
            if len(arg) != 0: arg = arg.rstrip(', ')
            
            v = eval("v.substitute(" + arg + ")")
            
            path.append(k)
            print("{} <--- {}".format("|".join( [str(x) for x in path] ), str(v)))
            
            nesting = ''
            for k in path: nesting += "['" + k + "']"
            paths_to_chg[nesting] = v
            path.pop()
            

        elif isinstance(v, list):
            for i in range(len(v)):
                if isinstance(v[i], obj):
                    
                    keys = list(set([ x[1] for x in Template.pattern.findall(v[i].template)]))
                    arg = ''
                    for key in keys:
                        arg += key + "=kw['" + key + "'], "
                    if len(arg) != 0: arg = arg.rstrip(', ')
                    
                    v[i] = eval("v[i].substitute(" + arg + ")")
                    path.append(k)
                    path.append(i)
                    print("{} <--- {}".format("|".join( [str(x) for x in path] ), str(v[i])))
                    
                    nesting = ''
                    for p in path: 
                        if isinstance(p, int): nesting += "[" + str(p) + "]"
                        else: nesting += "['" + p + "']"
            
                    paths_to_chg[nesting] = v[i]
                    path.pop()                    
                    path.pop()
        
        elif v == {}:
            pass
            #path.append(k)
            #path.pop()
            
        elif isinstance(v, dict):
            path.append(k)
            walk_for_obj(d=v, kw=kw, obj=obj)
            path.pop()

        elif isinstance(v, str):
            pass
                    
        elif isinstance(v, int):
            pass
        
        else:
            print("###Type {} not recognized: {}.{}={}".format(type(v), "|".join(path),k, v))

    return paths_to_chg


def parse_templates(dictionary, keywords):
    paths_to_chg = walk_for_obj(d=dictionary, kw=keywords, obj=string.Template)
    
    for k,v in paths_to_chg.items():
        #print(k,v)
        print( "dictionary" + k + ' = "' + v + '"')
        exec("dictionary" + k + ' = "' + v + '"',None,locals())

    return dictionary

# https://stackoverflow.com/questions/2220699/whats-the-difference-between-eval-exec-and-compile-in-python
# https://www.saltycrane.com/blog/2008/01/python-variable-scope-notes/
def eval_dict_values(dictionary):
    
    paths_to_chg = walk_for_functions(dictionary)
    
    for k,v in paths_to_chg.items():
#        print(k,v)
        print( "dictionary" + k + ' = eval("' + str(v) + '")')
        exec("dictionary" + k + ' = eval("' + str(v) + '")',None,locals())
    return dictionary

    
#==============================================================================
# generate pivot tables
#==============================================================================

#load templates for pivots and summary statistics and formatting
with open(config) as jdata:
    try:
        report = json.load(jdata) # np.sum will not be eval
    except ValueError as err:
        print(err)
        
        
# parse JSON inputs to evaluate objects and functions
report = eval_dict_values(report)
# parse Templates
report = parse_templates(report, keywords)

cfg_csv , cfg_excel = None, None
if '#input_config' in report.keys():
    if 'config_csv' in report['#input_config'].keys():
        cfg_csv = report['#input_config']['config_csv']
    if 'config_excel' in report['#input_config'].keys():
        cfg_excel = report['#input_config']['config_excel']

_cfg_csv , _cfg_excel = None, None
if not cfg_csv is None: _cfg_csv = pivot_table.config_csv(**cfg_csv)
if not cfg_excel is None: _cfg_excel = pivot_table.config_excel(**cfg_excel)


#create pivot tables
p = pivot_table.pivottable(filepath = src, dest_path = dest, \
                           csv_config = _cfg_csv, excel_config = _cfg_excel)
p.create_table(report)

# summary statistics on pivot table
if '#pivot_summary' in report.keys(): 
    summary = report['#pivot_summary']

    # calling functions with args from json input outside code
    for method, method_values in summary.items():
        
        arg=''
        if isinstance(method_values, dict):
            # multiple kwargs
            for kw, arg_value in method_values.items():
                arg += kw + '=' + '"' + str(arg_value) + '"' + ', '
            arg = arg.rstrip(', ')
            pass
        
        if isinstance(method_values, str):
            # only one arg, no kw
            arg = '"' + method_values + '"'
            pass
    
        # allow same functions stored as multiple keys
        if bool(re.match(r'^\d', method)): method = method.split("~")[1]
        print('p.' + method + '(' + arg + ')' )
        eval('p.' + method + '(' + arg + ')' )

# save computations from python to excel wb
p.closewb()


#==============================================================================
# format excel output
#==============================================================================

e = pivot_table.excel_editor(dest, visibility = False)
e.autofit_columns()
print(e.scan_columns())

# summary statistics on pivot table
if '#format_excel' in report.keys(): 
    steps = report['#format_excel']

    # calling functions with args from json input outside code
    for method, method_values in steps.items():
        
        arg=''
        if isinstance(method_values, dict):
            # multiple kwargs
            for kw, arg_value in method_values.items():
                if isinstance(arg_value, int) or isinstance(arg_value, float) or isinstance(arg_value, list):
                    arg += kw + '=' + str(arg_value) + ', '
                else:
                    arg += kw + '=' + "'" + str(arg_value) + "'" + ', '
            if len(arg) != 0: arg = arg.rstrip(', ')
            pass
        
        if isinstance(method_values, str):
            # only one arg, no kw
            if isinstance(arg_value, int) or isinstance(arg_value, float) or isinstance(arg_value, list):
                arg = method_values 
            else:
                arg = "'" + method_values + "'"
            pass
        
        # allow same functions stored as multiple keys
        if bool(re.match(r'^\d', method)): method = method.split("~")[1]
        print('e.' + method + '(' + arg + ')' )
        eval('e.' + method + '(' + arg + ')' )

e.close()








##==============================================================================
## 
##==============================================================================
#
##import chardet
##with open(filename, 'rb') as f:
##    result = chardet.detect(f.readline())  # or readline if the file is large
#
##Try calling read_csv with encoding='latin1', encoding='iso-8859-1' or encoding='cp1252'
#data = pd.read_csv(filename, encoding='iso-8859-1')
#



##############################################################3

{
 "#input_config":{"config_csv" : {"sep":"|", "header":0, "skipfooter":"None", "skiprows":"None", "encoding" : "ISO 8859-1",
					"usecols" : ["Client.Group","CG..Final.","BS_TYPE","Outstanding",
							"Mgt.Prod.Grouping","Standard.Product","Asset_Class","Exposure_for_RWA",
							"Country_of_Exposure","Orgnisational_unit_level_1",
							"CCR_Risk_weighted_calculation_method","STDF_Template_Indicator"]}
				},

"EA_total":{
	"val": "Exposure_for_RWA", 
	"idx": ["STDF_Template_Indicator", "Orgnisational_unit_level_1"], 
	"col": ["Reporting month"], 
	"filter": {}, 
	"aggfunc": "np.sum"}, 
"EA_product_group": {
	"val": "Exposure_for_RWA", 
	"idx": ["STDF_Template_Indicator", "Orgnisational_unit_level_1", "Mgt Prod Grouping"], 
	"col": ["Reporting month"], 
	"filter": {}, 
	"aggfunc": "np.sum"}, 
"EA_product": {
	"val": "Exposure_for_RWA", 
	"idx": ["STDF_Template_Indicator", "Orgnisational_unit_level_1", "Standard Product"], 
	"col": ["Reporting month"], 
	"filter": {}, 
	"aggfunc": "np.sum"}, 
"EA_asset_type": {
	"val": "Exposure_for_RWA", 
	"idx": ["STDF_Template_Indicator", "Orgnisational_unit_level_1", "Asset_Class"], 
	"col": ["Reporting month"], 
	"filter": {}, 
	"aggfunc": "np.sum"}, 
"EA_Country_of_Exposure": {
	"val": "Exposure_for_RWA", 
	"idx": ["STDF_Template_Indicator", "Orgnisational_unit_level_1", "Country_of_Exposure"], 
	"col": ["Reporting month"], 
	"filter": {}, 
	"aggfunc": "np.sum"}, 
"EA_client_group": {
	"val": "Exposure_for_RWA", 
	"idx": ["STDF_Template_Indicator", 
	"Client Group"], 
	"col": ["Reporting month"], 
	"filter": {}, 
	"aggfunc": "np.sum"}, 
"EA_credit_grade": {
	"val": "Exposure_for_RWA", 
	"idx": ["STDF_Template_Indicator", "CG Final"], 
	"col": ["Reporting month"], 
	"filter": {}, 
	"aggfunc": "np.sum"}, 
"Outstanding_total": {
	"val": "Outstanding", 
	"idx": ["STDF_Template_Indicator", "Orgnisational_unit_level_1"], 
	"col": ["Reporting month"], 
	"filter": {"BS_TYPE": ["ASS", "DAF", "DAO", "DLO", "DLF"]}, 
	"aggfunc": "np.sum"}, 
"Outstanding_product_group": {
	"val": "Outstanding", 
	"idx": ["STDF_Template_Indicator", "Orgnisational_unit_level_1", "Mgt Prod Grouping"], 
	"col": ["Reporting month"], 
	"filter": {"BS_TYPE": ["ASS", "DAF", "DAO", "DLO", "DLF"]}, 
	"aggfunc": "np.sum"}, 
"Outstanding_product": {
	"val": "Outstanding", 
	"idx": ["STDF_Template_Indicator", "Orgnisational_unit_level_1", "Standard Product"], 
	"col": ["Reporting month"], 
	"filter": {"BS_TYPE": ["ASS", "DAF", "DAO", "DLO", "DLF"]}, 
	"aggfunc": "np.sum"}, 
"Outstanding_client_group": {
	"val": "Outstanding", 
	"idx": ["STDF_Template_Indicator", "Client Group"], 
	"col": ["Reporting month"], 
	"filter": {"BS_TYPE": ["ASS", "DAF", "DAO", "DLO", "DLF"]}, 
	"aggfunc": "np.sum"}, 
"Outstanding_credit_grade": {
	"val": "Outstanding", 
	"idx": ["STDF_Template_Indicator", "CG Final"], 
	"col": ["Reporting month"], 
	"filter": {"BS_TYPE": ["ASS", "DAF", "DAO", "DLO", "DLF"]}, 
	"aggfunc": "np.sum"},
	
"#pivot_summary":{"diff_abs":{"pre":"Template('$pre')", "post":"Template('$post')"},
	"diff_perc":{"pre":"Template('$pre')", "post":"Template('$post')"},
	"sortPT" : "Template('$post - $pre (abs)')",
	"drop_col":"Total",
	"drop_row":"Total"},
	
"#format_excel":{
	"1~format_cells":{"which":["Template('$post - $pre (abs)')", "Template('$post')", "Template('$pre')"], "style":"Comma"},
	"2~format_cells":{"which": ["Template('$post - $pre (%)')"], "style":"Percent", "decimal_places":5}}	

}

#########################################3


import win32com.client
#from win32com.client.gencache import EnsureDispatch as Dispatch
import pandas as pd
import re
import os

#==============================================================================
# Read current Window's Outlook Global Address List
#==============================================================================

class outlookGAL(object):

    def __init__(self):
        self._outlook = win32com.client.gencache.EnsureDispatch('Outlook.Application')
        self._BankID = ''
        self.results = {}

    def _getfromOutlook(self, BankID, isManager = False, workerBankID = None):
        
        recipient = self._outlook.Session.CreateRecipient(BankID)
        rev = recipient.Resolve()
        if not rev:
            print("BankID: " + str(BankID) + " is not found in Global Address List!")
            return
        else:
            ae = recipient.AddressEntry
            u = ae.GetExchangeUser()
            
        if 'EX' == ae.Type: email_address = u.PrimarySmtpAddress
        if 'SMTP' == ae.Type: email_address = ae.Address
        #https://social.msdn.microsoft.com/Forums/expression/en-US/9faa0862-4824-4691-8531-fe403a7eb3ff/how-can-i-addgo-back-to-150-office-library-references-after-2016-update?forum=accessdev
        #print(recipient.Name + ' is ' + 'sendable' if recipient.Sendable else 'not sendable')
        
        if BankID not in list(self.results.keys()):
            self.results[BankID] = {}
            self.results[BankID]['BankID'] = BankID
            self.results[BankID]['Name'] = recipient.Name 
            self.results[BankID]['Department'] = u.Department
            self.results[BankID]['JobTitle'] = u.JobTitle
            self.results[BankID]['CompanyName'] = u.CompanyName
            self.results[BankID]['City'] = u.City
            self.results[BankID]['StreetAddress'] = u.StreetAddress
            self.results[BankID]['OfficeLocation'] = u.OfficeLocation
            self.results[BankID]['MobileTelephoneNumber'] = u.MobileTelephoneNumber
            self.results[BankID]['email'] = email_address
        
        if isManager:
            try:
                self.results[BankID]['workerBankID']
            except KeyError:
                self.results[BankID]['workerBankID'] = []
            
            self.results[BankID]['workerBankID'].append(workerBankID)
        
        for k,v in self.results[BankID].items():
            print(str(k) + ": " + str(v))
        print("==================================\n")

    def find(self, *args, **kwds):
        if(len(args)==1 and isinstance(args[0],str)): 
            self._BankID = args[0]
            self._getfromOutlook(self._BankID)
        elif(len(args)==1 and isinstance(args[0],list)): 
            list_BankID = args[0]
            for BankID in list_BankID: self.find(BankID)
        else:
            print("BankID needs to be in a list or a str")
            
    def find_manager(self, *args, **kwds):
        if(len(args)==1 and isinstance(args[0],str)):
            BankID = args[0]
        
        recipient = self._outlook.Session.CreateRecipient(BankID)
        rev = recipient.Resolve()
        if not rev: 
            print("BankID: " + str(BankID) + " is not found in Global Address List!")
            return
        else:
            ae = recipient.AddressEntry
            u = ae.GetExchangeUser()
            m = u.GetExchangeUserManager()
            # Manager's BankID
            try: 
                m_BankID = int(m.Address[ m.Address.rfind("/cn=")+4 : m.Address.rfind("/cn=")+4+7])
            except ValueError:
                print('Manager BankID is special in Address')
                m_BankID = m.Address[ m.Address.rfind("/cn=")+4 : len(m.Address)-3]
            except AttributeError:
                print('Invalid BankID!')
                return
                
            print("Manager's BankID: " + str(m_BankID) if not isinstance(m_BankID,str) else str(m_BankID) )
            self._getfromOutlook(str(m_BankID), isManager = True, workerBankID = str(BankID))
        
    def find_orgchart(self, *args, **kwds):
        if(len(args)==1 and isinstance(args[0],str)): 
            BankID = args[0]

        recipient = self._outlook.Session.CreateRecipient(BankID)
        rev = recipient.Resolve()
        if not rev: 
            print("BankID: " + BankID + " is not found in Global Address List!")
            return
        
        prevBankID = ''
        while BankID != prevBankID:
            u = self._outlook.Session.CreateRecipient(BankID).AddressEntry.GetExchangeUser()
            print(u.Name)
            prevBankID = BankID
            BankID = u.GetExchangeUserManager().Name

    def find_all_managers(self, *args, **kwds):
        if(len(args)==1 and isinstance(args[0],list)): 
            list_BankID = args[0]
            for BankID in list_BankID: self.find_manager(BankID)
        elif(len(args)==1 and isinstance(args[0],str)): 
            list_BankID = args[0]
            self.find_manager(list_BankID)
        else:
            print("BankID: " + BankID + " is not found in Global Address List!")
            
    def get_results(self):
        return pd.DataFrame(self.results).transpose()
    
    def del_results(self):
        self.results = {}
            
    __doc__ = """
    ======================================================
    USES
    ======================================================
    - Find contact details from Outlook Global Address List
    
    ======================================================
    SAMPLE
    ======================================================
    gal = outlookGAL()
    gal.find('1522918')
    gal.find_manager('1522918')
    gal.find('Winters, Bill')
    gal.find_manager('Winters, Bill')
    gal.find_manager('Vinals, Jose')
    gal.find_orgchart('1580873')
    gal.results
    
    gal = outlookGAL()
    list_BankID = ['1289066','1379266','1216415','1580873']
    gal.find_all(list_BankID)
    gal.results
    b=gal.get_results()
    """

#==============================================================================
# 
#==============================================================================

class FileWriter(object):
    
    def __init__(self, filename):
        self.file = open(filename, "w")

    def printlog(self, a_string):
        str_uni = a_string.encode('utf-8')
        self.file.write(str(str_uni))
        self.file.write("\n")
        print(a_string)

    def flush(self):
        self.file.flush()
    
    __doc__ = """
    ====================================================================
    USES
    ====================================================================
    - print string from python to log.txt saved on local drive
    ====================================================================    
    SAMPLE
    ====================================================================
    log = FileWriter("filename.txt")
    log.printlog("some string to log into .txt file")
    log.flush()
    """
        
        
class CheckMail(object):
    
    def __init__(self, email, mailbox, startDate=None, endDate=None, all_msgs = False): #startDate, endDate):
        self._log = FileWriter(re.sub("\.","",email[:email.find('@')]) + "_" + mailbox + ".txt")
        self._outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        self._account = self._outlook.Folders[email]
        self._mailbox = self._account.Folders[mailbox]
        self.results = [] # store log in dict {subject: body}
        self._startDate = startDate
        self._endDate = endDate
        self._all_msgs = all_msgs
        
        # check for dates
        try:
            self._startDate
            self._endDate
        except (NameError, UnboundLocalError) as e:
            print('Please enter in Start and End Date since you are not reading all messages!')

        if not isinstance(self._startDate, pd.Timestamp): self._startDate = pd.to_datetime(startDate)
        if not isinstance(self._endDate, pd.Timestamp): self._endDate = pd.to_datetime(endDate)
    
    
    def _logger(self, msg):
        #==============================================================================
        # Microsoft Outlook 16.0 Object Library Methods
        
        #mailbox="JustinSHuiMing.Yeo@sc.com"
        #folderindex = 'Inbox'
        #msgs=win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI").Folders[mailbox].Folders[folderindex].Items
        #msg = msgs.GetLast()

        #msg.Subject
        #msg.Body
        #msg.SenderName
        #msg.ReceivedTime
        
        #msg.Attachments.Count
        #for i in range(1, msg.Attachments.Count+1 ):
        #    msg.Attachments.Item(i).DisplayName
        #==============================================================================
        self._log.printlog("Message Received Time at: " + str(msg.ReceivedTime))
        self._log.printlog("Sender: " + msg.SenderName)
        self._log.printlog("Subject: " + msg.Subject)
        self._log.printlog("Body: " + msg.Body)
        
        print("Checking on... " + msg.Subject)
        
        atch = [] # attachment names
        if msg.Attachments.Count >=1 :
            for i in range(1,msg.Attachments.Count+1) :
                self._log.printlog("Attachment " + str(i) + " : " + msg.Attachments.Item(i).DisplayName)
                atch.append(msg.Attachments.Item(i).DisplayName)
        
        # save for return
        self.results.append({"Subject": msg.Subject, 
                             "Time": msg.ReceivedTime,
                             "Sender": msg.SenderName,
                             "Body": msg.Body,
                             "Attachments": atch
                             })
    
    def check(self):
        
        # initialise HEAD
        msgs = self._mailbox.Items
        try: 
            msg = msgs.GetLast() # most recent
        except: 
            print(mailbox + ' has no mails!')
            
        while msg:
            # limit to received DateTime to (startDate, endDate)
            if self._all_msgs == False: 
                # stop if beyond filter
                if msg.ReceivedTime.date() > self._endDate.date():
                    msg = msgs.GetPrevious()
                    continue
                if msg.ReceivedTime.date() < self._startDate.date(): 
                    break             
            
            self._logger(msg)
            msg = msgs.GetPrevious()
        
        self._log.printlog("-" * 50)
        self._log.flush()
        
        
    def find(self, pattern, where):
        if where=='Attachments': print("use .find_attachments for Attachments. Only for Body, Subject,... ")
        # find for pattern = specific keywords + variant from where = Subject, Body, ...             
        pattern = pattern.lower()
        hasit = [ True if pattern in m[where].lower() else False for m in self.results ]
        return hasit
    
    def get(self,what):
        return [x[what] for x in self.results]
    
#    def find_attachments(self, pattern, download=False):
#        att = self.get('Attachments')
#        cwd = os.getcwd()
#        hasit = [False for x in range(len(att))]
#        for j in range(len(att)):
#            if len(att[j])>=1:
#                for a in att[j]:
#                    if pattern.lower() in a.lower(): 
#                        hasit[j]
#                        # if need to download
#                        if download:
#                            cwd + "\\" + 
#                        continue
#        return hasit
    
    __doc__ = """
    ====================================================================
    USES
    ====================================================================
    - Find in subject, body, sendername whether it contains a string pattern
    ====================================================================    
    SAMPLE
    ====================================================================
    mail = CheckMail(email = email, mailbox = mailbox, all_msgs=True)
    mail = CheckMail(email = email, mailbox = mailbox, all_msgs=False, startDate = '20180105', endDate = '20180117')
    
    mail.check()
    
    mail.results
    len(mail.results)
    tf=mail.find(pattern = 'justin', where = 'Subject')
    list(map(mail.results.__getitem__, [i for i in range(len(tf)) if tf[i] is True] ))
    tf=mail.find(pattern = 'justin', where = 'Body')
    list(map(mail.results.__getitem__, [i for i in range(len(tf)) if tf[i] is True] ))
    for i in list(map(mail.results.__getitem__, [i for i in range(len(tf)) if tf[i] is True] )):
        print(i['Body'])
    """


    
   ###############################


import win32com.client
import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil

#os.chdir('C:\\Users\\1580873\\Desktop\\RiskView')
import outlook_helpers
#from importlib import reload
#reload(outlook_helpers)

#==============================================================================
# handler class for incoming mail (FOR pythoncom LISTENING)
#==============================================================================

class Handler_Class(object, ):
    def __init__(self):
        self._outlook = win32com.client.Dispatch('outlook.application')
        self._attachment = None
        self._GAL = outlook_helpers.outlookGAL()
        self._results = {}
        
    def OnNewMailEx(self, receivedItemsIDs):
        # https://msdn.microsoft.com/en-us/vba/outlook-vba/articles/application-newmailex-event-outlook
        for ID in receivedItemsIDs.split(","):
            # multiple mails incoming simultaneously
            mailItem = self._outlook.Session.GetItemFromID(ID)
            print("Subject: " + mailItem.Subject)
            print("Body: " + mailItem.Body)
            print("Sender: " + mailItem.SenderName)
#            print("att counts" + str(mailItem.Attachments.Count))
#            print('attach:' + mailItem.Attachments.Item(1).DisplayName)

            if 'autoreply' in mailItem.Subject.lower():
                print("Don't reply an auto reply...")
                return
            elif 'riskview' in mailItem.Subject.lower(): request_type = 'RiskView'
            elif 'ccbr' in mailItem.Subject.lower(): request_type = 'CCBR'
            #elif 'mtcr' in mailItem.Subject.lower(): request_type = 'MTCR'
            else: 
                print('Not a mail to reply... No action!')
                return
            
            switcher = {
                    'RiskView': self._OnNewMailEx_RiskView,
                    'CCBR': self._OnNewMailEx_CCBRView #,'MTCR': self._OnNewMailEx_MTCRView
                        }
    
            # case: call private method
            process = switcher.get(request_type, lambda: "invalid mail checker rule!")
            process(mailItem)
            print("===============================")
    
    def _OnNewMailEx_RiskView(self, mailItem):
        
        check = self._OnNewMail_RiskView_filter(mailItem)
        if check is True:
            # edit data into exisiting file                
            # key --> RequestType|BankID|Name|Department
            fields = list(self._results.keys())[0].split('|')
            
            record = pd.DataFrame(fields[1:]+[None]).transpose()
            record.columns = ['BankID','Name','UserGroup','Dummy']
            
            rw_subsegmentaccess = self._results[list(self._results.keys())[0]]['rw_subsegmentaccess']
            subsegmentcode = rw_subsegmentaccess.loc[rw_subsegmentaccess['Grant Access']=='Yes','SubSegment Code'].to_frame()                
            subsegmentcode['Dummy'] = None
            subsegmentcode.columns = ['Subsegment','Dummy']
            
            rw_countryaccess = self._results[list(self._results.keys())[0]]['rw_countryaccess']
            countrycode = rw_countryaccess.loc[rw_countryaccess['Grant Access']=='Yes','Country Code'].to_frame()                
            countrycode['Dummy'] = None
            countrycode.columns = ['CountryCode','Dummy']
            
            all_CB = all([ x == 'Yes' for x in rw_subsegmentaccess.loc[rw_subsegmentaccess['CB / CIB']=='CB','Grant Access'] ])
            all_CIB = all([ x == 'Yes' for x in rw_subsegmentaccess.loc[rw_subsegmentaccess['CB / CIB']=='CIB','Grant Access'] ])
            all_Subsegment = all([ x == 'Yes' for x in rw_subsegmentaccess.loc[:,'Grant Access'] ])
            global_Countries = all([ x == 'Yes' for x in rw_countryaccess.loc[:,'Grant Access'] ])

            if global_Countries and \
                (all_CB or all_CIB or all_Subsegment):
                # Global + CB / Global + CIB / Global + all subsegment
                # copy to DASHBOARD access only, not for RiskView Access / CCBRView Access
                # can keep a seperate internal sheet indpt from Tableau for recordkeeping of global access users
                d_tocopy = record.drop('Dummy',axis=1)
                d_tocopy['Project Access'] = 'RiskView'
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'DashboardAccess', \
                        datatoappend = d_tocopy)
                
                d_tocopy = d_tocopy.drop('Project Access',axis=1)
                
                if all_Subsegment:                    
                    d_tocopy['Global Type'] = "Global/Global"
                elif all_CB:
                    d_tocopy['Global Type'] = "Global/CB"
                elif all_CIB:
                    d_tocopy['Global Type'] = "Global/CIB"
                
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'RiskViewGlobalViewAccess', \
                        datatoappend = d_tocopy)
                
            else:
                #custom countries
                # CROSS APPLY if custom
                tocopy = pd.merge(record, subsegmentcode, on='Dummy')
                tocopy = pd.merge(tocopy, countrycode, on='Dummy').drop('Dummy',axis=1)
                # copy and paste into results or data of form into Pending_Access.xlsx Stack
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                                        sheetname = 'RiskViewAccess', \
                                        datatoappend = tocopy)
                d_tocopy = record.drop('Dummy',axis=1)
                d_tocopy['Project Access'] = 'RiskView'
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'DashboardAccess', \
                        datatoappend = d_tocopy)
            
            # save mailItem as .msg into cwd\\folder            
            if not os.path.isdir(os.getcwd() + '\\' + 'RiskView_MailItems'):
                os.mkdir('RiskView_MailItems')
            msgName = 'RiskView Request ' + fields[0]
            mailItem.SaveAs(os.getcwd() + '\\' + 'RiskView_MailItems\\' + msgName + '.msg')

        else:
            # failed filter, auto reply rejection
            to = mailItem.SenderName
            subject = "AUTOREPLY: " + mailItem.Subject
            body = self._rejection_pretty(check)
            self._send_mail(to, subject, body, mailItem) # handles 0 or multiple attachments

        pass
        
    def _OnNewMailEx_CCBRView(self, mailItem):
        
        check = self._OnNewMail_CCBRView_filter(mailItem, approver_department = 'Group Country Risk')
        if check is True:
            # edit data into exisiting file                
            # key --> RequestType|BankID|Name|Department
            fields = list(self._results.keys())[0].split('|')
            
            record = pd.DataFrame(fields[1:]+[None]).transpose()
            record.columns = ['BankID','Name','UserGroup','Dummy']
            
            rw_countryaccess = self._results[list(self._results.keys())[0]]['rw_countryaccess']
            countrycode = rw_countryaccess.loc[rw_countryaccess['Grant Access']=='Yes','Country Code'].to_frame()                
            countrycode['Dummy'] = None
            countrycode.columns = ['CountryCode','Dummy']
            
            global_Countries = all([ x == 'Yes' for x in rw_countryaccess.loc[:,'Grant Access'] ])

            if global_Countries:
                # Global + CB / Global + CIB / Global + all subsegment
                # copy to DASHBOARD access only, not for RiskView Access / CCBRView Access
                # can keep a seperate internal sheet indpt from Tableau for recordkeeping of global access users
                d_tocopy = record.drop('Dummy',axis=1)
                d_tocopy['Project Access'] = 'CCBRView'
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'DashboardAccess', \
                        datatoappend = d_tocopy)

                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'CCBRViewGlobalViewAccess', \
                        datatoappend = d_tocopy.drop('Project Access',axis=1))
                
            else:
                #custom countries
                # CROSS APPLY if custom
                tocopy = pd.merge(record, countrycode, on='Dummy').drop('Dummy',axis=1)
                # copy and paste into results or data of form into Pending_Access.xlsx Stack
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                                        sheetname = 'CCBRViewAccess', \
                                        datatoappend = tocopy)
                d_tocopy = record.drop('Dummy',axis=1)
                d_tocopy['Project Access'] = 'CCBRView'
                self._append_to_pending(filename = 'Pending_Access.xlsx', \
                        sheetname = 'DashboardAccess', \
                        datatoappend = d_tocopy)
            
            # save mailItem as .msg into cwd\\folder            
            if not os.path.isdir(os.getcwd() + '\\' + 'RiskView_MailItems'):
                os.mkdir('RiskView_MailItems')
            msgName = 'RiskView Request ' + fields[0]
            mailItem.SaveAs(os.getcwd() + '\\' + 'RiskView_MailItems\\' + msgName + '.msg')
            
        else:
            # failed filter, auto reply rejection
            to = mailItem.SenderName
            subject = "AUTOREPLY: " + mailItem.Subject
            body = self._rejection_pretty(check)
            self._send_mail(to, subject, body, mailItem) # handles 0 or multiple attachments

        pass
        
    def _OnNewMailEx_MTCRView(self, mailItem):
        
        check = self._OnNewMail_MTCRView_filter(mailItem)
        if check is True:
            pass
        else:
            # failed filter, auto reply rejection
            pass
    
    def _append_to_pending(self, filename, sheetname, datatoappend, index=False, header=False):
        #with open('Pending_Access.xlsx', 'a') as f:
            #    df.to_csv(f, header=True, index_label=False)
            
        # http://openpyxl.readthedocs.io/en/default/pandas.html
        
        # from openpyxl import load_workbook
        # from openpyxl.utils.dataframe import dataframe_to_rows            
        wb = load_workbook(filename = filename)
        #wb.sheetnames
        sht = wb[sheetname]
        #sht.max_row
        for row in dataframe_to_rows(datatoappend, index=index, header=header):
            sht.append(row)
        
        wb.save(filename)
        wb.close()
        del(row)
        del(sht)
        del(wb)        
        pass
    
    
    def _OnNewMail_CCBRView_filter(self, mailItem, *args, **kwargs):
        '''
        *args, **kwargs
        approver_department = "Group Country Risk"
        '''
        # CORRECT EMAIL SUBJECT
        rule1 = 'request' or 'ccbr' in str(mailItem.Subject).lower()
        if not rule1: return "Mail subject incorrectly named.\n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        # HAS ATTACHMENT AND ONLY RELEVANT ONE ATTACHMENT
        rule2 = mailItem.Attachments.Count == 1
        if not rule2: return "Mail do not have the request form attachment and can only have one attachment.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION
        
        # CHECK FILENAME OF ATTACHMENT FORM
        rule3 = mailItem.Attachments.Item(1).DisplayName == "CCBRView Access Request Form.xlsx"
        if not rule3: return "Mail do not have the correct request form attachment.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION


        if not os.path.isdir(os.getcwd() + '\\' + 'temporary_for_attachments'):
            os.mkdir('temporary_for_attachments')
        old_dir = os.getcwd()
        os.chdir(os.getcwd() + '\\' + 'temporary_for_attachments')
        # OPEN ATTACHMENT
        self._open_attachments(mailItem, whichItem=1, sheetname = 0, 
                               index_col=None, header=None, savein = os.getcwd())
        os.chdir(old_dir)
        #os.rmdir('temporary_for_attachments')
        
        # TALLY BankID TO ENSURE macros was used to send request form
        rule4 = len(re.findall(r"(?<!\d)\d{7}(?!\d)", mailItem.Subject)) == 1
        if not rule4: return "Request form can only have one requester's BankID.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION

        requester_BankID = str(self._attachment.iloc[0,1])
        rule5 = requester_BankID == str(re.findall(r"(?<!\d)\d{7}(?!\d)", mailItem.Subject)[0])
        if not rule5: return "Please include requester's BankID in mail subject.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION
        
        # VALIDATE DATE ACCURACY AND COMPLETENESS IN REQUEST FORM
        rw_countryaccess = self._attachment.iloc[4:265, 0:4]
        rw_countryaccess.columns = rw_countryaccess.iloc[0]
        rw_countryaccess = rw_countryaccess.reindex(rw_countryaccess.index.drop(4))
        rule6 = rw_countryaccess.shape == (260, 4) and \
                all(entry in ['Yes','No'] for entry in rw_countryaccess['Grant Access'])
        if not rule6: return "Request form attachment country access data is invalidated.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION
        
        # So far, SENDER is anyone from Department Group Country Risk
        # requester --> line manager --> our inbox
        
        self._GAL.find(requester_BankID)
        try:
            rule8 = self._GAL.results[requester_BankID]["Department"] == kwargs['approver_department']
        except:
            print('Key error: CCBRView filter needs an approver department from ' + kwargs['approver_department'])
            return        
        
        self._GAL.del_results()
        
        if not rule8: return "Warning: Approver for CCBRView should be from the Group Country Risk Department.\n\
        Please get someone from the Group Country Risk to approve it." # proceed to AUTO REJECTION
                
        # passed all rejection filters
        # key --> RequestType|BankID|Name|Department
        self._GAL.find(requester_BankID)
        name = self._GAL.get_results()['Name'].iloc[0]
        dept = self._GAL.get_results()['Department'].iloc[0]
        self._GAL.del_results()
        
        self._results["CCBRView|"+str(requester_BankID)+"|"+str(name)+"|"+str(dept)] = \
                        {"rw_countryaccess":rw_countryaccess,
                         "Business Justification":self._attachment.iloc[1,1]}
        return True

    def _OnNewMail_MTCRView_filter(self, mailItem):
        # So far, SENDER is Head of MTCR # approver = "Poquet, Morgan"
        # requester --> line manager --> our inbox
        approver = mailItem.SenderName
        self._GAL.find(approver)
        rule1 = self._GAL.results[approver]["JobTitle"] in \
            ["Head, MTCR - Traded Credit & Credit Trading", "Global Head Market & Traded Credit Risk"]
        self._GAL.del_results()
        
        # Alternatively, SENDER is Head of MTCR is part of FWD mail thread
        # requester --> line manager --> requester FWD --> our inbox
        if not rule1:
            single_mail_body = self._get_single_email_thread(mailItem.Body, approver)
            mail_list = [x.strip(',.-><:;') for x in single_mail_body.lower().split()]
            matchingkw = ['approve']
            rule1 = any(x in matchingkw for x in mail_list)
        
        if not rule1: return "Please be informed that only the Head of MTCR can approve request access.\n\
        Please get approval from the Head of MTCR and resubmit." # proceed to AUTO REJECTION    
        
        # passed all rejection filters
#        self._results["MTCRView|"+str(requester_BankID)+"|"+str(name)+"|"+str(dept)] = \
#                        {"rw_countryaccess":rw_countryaccess,
#                         "Business Justification":self._attachment.iloc[1,1]}
        return True
        
    def _OnNewMail_RiskView_filter(self, mailItem, *args, **kwargs):
        '''
        *args, **kwargs
        matchingkw = ['approve']
        '''
        # CORRECT EMAIL SUBJECT
        rule1 = 'request' and 'riskview' in str(mailItem.Subject).lower()
        if not rule1: return "Mail subject incorrectly named. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        # HAS ATTACHMENT AND ONLY RELEVANT ONE ATTACHMENT
        rule2 = mailItem.Attachments.Count == 1
        if not rule2: return "Mail do not have the request form attachment and can only have one attachment. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        # CHECK FILENAME OF ATTACHMENT FORM
        rule3 = mailItem.Attachments.Item(1).DisplayName == "RiskView Access Request Form.xlsm"
        if not rule3: return "Mail do not have the correct request form attachment. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        if not os.path.isdir(os.getcwd() + '\\' + 'temporary_for_attachments'):
            os.mkdir('temporary_for_attachments')
        old_dir = os.getcwd()
        print(old_dir)
        os.chdir(os.getcwd() + '\\' + 'temporary_for_attachments')
        # OPEN ATTACHMENT
        self._open_attachments(mailItem, whichItem=1, sheetname = 0, 
                               index_col=None, header=None, savein = os.getcwd())
        os.chdir(old_dir)
        shutil.rmtree(os.getcwd() + '//' + 'temporary_for_attachments')
        
        # TALLY BankID TO ENSURE macros was used to send request form
        rule4 = len(re.findall(r"(?<!\d)\d{7}(?!\d)", mailItem.Subject)) == 1
        if not rule4: return "Request form can only have one requester's BankID. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION        

        requester_BankID = str(self._attachment.iloc[0,1])
        rule5 = requester_BankID == str(re.findall(r"(?<!\d)\d{7}(?!\d)", mailItem.Subject)[0])
        if not rule5: return "Please include requester's BankID in mail subject.\n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        # VALIDATE DATE ACCURACY AND COMPLETENESS IN REQUEST FORM
        rw_countryaccess = self._attachment.iloc[4:263, 0:4]
        rw_countryaccess.columns = rw_countryaccess.iloc[0]
        rw_countryaccess = rw_countryaccess.reindex(rw_countryaccess.index.drop(4))
        rule6 = rw_countryaccess.shape == (258, 4) and \
                all(entry in ['Yes','No'] for entry in rw_countryaccess['Grant Access'])
        if not rule6: return "Request form attachment country access data is invalidated. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION
        
        rw_subsegmentaccess = self._attachment.iloc[4:48, 5:11]
        rw_subsegmentaccess.columns = rw_subsegmentaccess.iloc[0]
        rw_subsegmentaccess = rw_subsegmentaccess.reindex(rw_subsegmentaccess.index.drop(4))
        rule7 = rw_subsegmentaccess.shape == (43, 6) and \
                all(entry in ['Yes','No'] for entry in rw_subsegmentaccess['Grant Access'])
        if not rule7: return "Request form attachment subsegment access data is invalidated. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION

        # CHECK APPROVER who send the email is the requester's LINE MANAGER
        # requester --> line manager --> our inbox
        self._GAL.find_manager(requester_BankID)
        if len(self._GAL.results) == 1: # one worker --> one line manager
            approver_BankID = list(self._GAL.results.keys())[0]
            approver_name = self._GAL.results[approver_BankID]['Name']
            sendername = mailItem.SenderName # line manager needs to send the mail via macros
            rule8 = sendername == approver_name
        else:
            rule8 = False # no manager, invalid BankID
        self._GAL.del_results()
        
        # ALTERNATIVE CHECK OF APPROVER (those who did not use the macros)
        # requester --> line manager --> requester FWD --> our inbox
        if not rule8:
            single_mail_body = self._get_single_email_thread(mailItem.Body, approver_name)
            mail_list = [x.lower().strip(',.-><:;') for x in single_mail_body]
            matchingkw = ['approve']
            rule8 = any(x in matchingkw for x in mail_list)

        if not rule8: return "Warning: Approver is not the requester's direct line manager. \n\
        Please get the requester's line manager to send the mail with approval using the macros." # proceed to AUTO REJECTION          
        
        # passed all rejection filters
        # key --> RequestType|BankID|Name|Department
        self._GAL.find(requester_BankID)
        name = self._GAL.get_results()['Name'].iloc[0]
        dept = self._GAL.get_results()['Department'].iloc[0]
        self._GAL.del_results()
        
        self._results["RiskView|"+str(requester_BankID)+"|"+str(name)+"|"+str(dept)] = \
                        {"rw_countryaccess":rw_countryaccess,
                         "rw_subsegmentaccess":rw_subsegmentaccess,
                         "Business Justification":self._attachment.iloc[1,1]}
        
        return True

    def _get_single_email_thread(self, mailItemBody, from_who):
        # If in Body: Must be from: <Line Manager Name> and below it includes Approved before next From
        # found in: SenderName, Subject, Body, 
        # extra mail Attributes: ReceivedTime, Attachments
        
        pattern = "From: " + from_who
        single_mail = mailItemBody[mailItemBody.find(pattern):]
        ending = single_mail[len(pattern):].find('From: ')
        ending = None if ending == -1 else ending + len(pattern)
        
        if ending is None:
            if single_mail == '\n': return []
            return [single_mail]
        else:
            return [single_mail[ :ending]] + \
                self._get_single_email_thread(self, single_mail[ending: ], from_who)


    def _rejection_pretty(self, msg):
        # rejection auto reply
        return "================================" + "\n" + \
        " ROBOT AUTO REJECTION " + "\n" + \
        "================================" + "\n" + \
        msg + "\n" + "\n" + \
        "Please resubmit with the relevant attachment. Thanks for your cooperation!" + "\n" + \
        "\n" + \
        "From the Credit Risk Monitoring Team"

    def _save_attachments(self, mailItem, whichItem=1, savein = os.getcwd()):
        '''
        args=['io', 'sheetname', 'header', 'skiprows', 'skip_footer', 'index_col', 
        'names', 'parse_cols', 'parse_dates', 'date_parser', 'na_values', 'thousands', 
        'convert_float', 'has_index_names', 'converters', 'dtype', 'true_values', 
        'false_values', 'engine', 'squeeze'], varargs=None, keywords='kwds', 
        '''
        # .xlsm means cannot copy paste into another excel wb
        # Don't accept i.e. Copy of RiskView Access Request Form.xlsm
        filename = mailItem.Attachments.Item(whichItem).DisplayName
        # save current attachment as temp in local drive
        print("saving attachments in... " + savein + '\\' + str(filename))
        mailItem.Attachments.Item(whichItem).SaveAsFile(savein + '\\' + str(filename))
        return savein + '\\' + str(filename)

    def _open_attachments(self, mailItem, whichItem=1, savein = os.getcwd(), 
                          *args, **kwargs):
        filename = self._save_attachments(mailItem, whichItem, savein)
        # read from local drive for processing
        self._attachment = pd.read_excel(io = filename, *args, **kwargs)
        pass
    
    def _send_mail(self, to, subject=None, body=None, attachments=None):
        mail = self._outlook.CreateItem(0)
        mail.To = to
        #Msg.CC = "more email addresses here"
        #Msg.BCC = "more email addresses here"
        if subject is not None: mail.Subject = subject
        if body is not None: mail.Body = body
        #In case you want to attach a file to the email
        #attachment  = "C:\\Users\\1580873\\Desktop\\IFRS9\\Script\\anthony_stdf_mapping.R"
        
        if not os.path.isdir(os.getcwd() + '\\' + 'temporary_for_attachments'):
            os.mkdir('temporary_for_attachments')
        old_dir = os.getcwd()
        os.chdir(os.getcwd() + '\\' + 'temporary_for_attachments')
        if attachments is not None:
            if isinstance(attachments,str): 
                mail.Attachments.Add(attachments)
            # handle multiple attachments ['C://...', 'C://...']
            elif isinstance(attachments,list): 
                for att in attachments:
                    mail.Attachments.Add(att)
            # forward attachments from another Outlook.mailItem
            elif 'win32com.gen_py' and 'MailItem' in str(type(attachments)):
                for i in range(1,1+attachments.Attachments.Count):
                    filename = self._save_attachments(attachments, whichItem = i, savein = os.getcwd())
                    mail.Attachments.Add(filename)
                    os.remove(filename)
            else:
                attachments = None
        mail.Send()
        os.chdir(old_dir)
        shutil.rmtree(os.getcwd() + '//' + 'temporary_for_attachments')
        pass
    
    
    __doc__ = """
    ====================================================================
    USES
    ====================================================================
    - print string from python to log.txt saved on local drive
    ====================================================================    
    SAMPLE
    ====================================================================
    log = FileWriter("filename.txt")
    log.printlog("some string to log into .txt file")
    log.flush()
    """            



    
  #######################################






























